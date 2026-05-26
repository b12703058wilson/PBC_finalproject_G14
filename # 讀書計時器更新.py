# 讀書計時器 — 路線分歧版
# 新增：故事線選擇（乙女向 / 恐怖向）、劇情分歧系統
# STORIES 改為依 route 分組的 dict；序章為共用，選完 route 後才分流

import subprocess
import sys
from dataclasses import dataclass, field

required_packages = ["openpyxl", "matplotlib"]
for package in required_packages:
    try:
        __import__(package)
    except ImportError:
        print(f"安裝 {package} 中，請稍等...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", package])

import tkinter as tk
import time
import os
import math
import datetime
import random
from openpyxl import Workbook, load_workbook
import matplotlib
matplotlib.use("TkAgg")
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg


# ════════════════════════════════════════════════
#  設定介面、按鈕、文字的顏色(方便之後做介面的時候用)
# ════════════════════════════════════════════════

C = {
    "bg":         "#FAFAF8",
    "surface":    "#FFFFFF",
    "surface2":   "#F5F3EE",
    "border":     "#E8E4DE",
    "border2":    "#D5CFC6",
    "pink":       "#D4537E",
    "pink_light": "#ED93B1",
    "pink_pale":  "#FBEAF0",
    "pink_text":  "#993556",
    "teal":       "#1D9E75",
    "teal_pale":  "#E1F5EE",
    "teal_text":  "#0F6E56",
    "gold":       "#C8922A",
    "gold_pale":  "#FDF3E0",
    "gold_bar":   "#E8B84B",
    "gray":       "#888780",
    "gray_pale":  "#F1EFE8",
    "gray_text":  "#5F5E5A",
    "text":       "#2C2C2A",
    "text2":      "#5F5E5A",
    "text3":      "#888780",
    "bubble_char":"#FFF0F5",
    "horror_bg":  "#1A1A2E",       # 恐怖向故事背景
    "horror_text":"#E8D5B7",       # 恐怖向故事文字
    "horror_btn": "#2D1B4E",       # 恐怖向按鈕背景
}

W, H = 400, 680


# ════════════════════════════════════════════════
#  路線設定
# ════════════════════════════════════════════════

ROUTE_CONFIG = {
    "romance": {
        "label":       "🌸  乙女向",
        "description": "與角色漸漸拉近距離。\n透過每次的讀書，解鎖更多故事。",
        "btn_bg":      "#D4537E",
        "btn_fg":      "#FFFFFF",
        "story_bg":    "#FFF9F3",
        "story_fg":    "#4A1B0C",
    },
    "horror": {
        "label":       "🕯  恐怖向",
        "description": "有些東西不該被發現。\n繼續讀下去，如果你敢的話。",
        "btn_bg":      "#2C2C3E",
        "btn_fg":      "#E8D5B7",
        "story_bg":    "#1A1A2E",
        "story_fg":    "#E8D5B7",
    },
}


# ════════════════════════════════════════════════
#  好感度階層
# ════════════════════════════════════════════════

# 設定好感度門檻(日常劇情使用)
AFFECTION_LEVELS = [
    {"threshold": 0,  "name": "陌生", "desc": "對方對你還不太熟悉……"},
    {"threshold": 20, "name": "禮貌", "desc": "對方以禮相待，但仍保持距離。"},
    {"threshold": 40, "name": "親近", "desc": "對方開始對你展露笑容。"},
    {"threshold": 60, "name": "信任", "desc": "對方願意和你分享心事了。"},
    {"threshold": 80, "name": "親密", "desc": "對方在你身邊感到很自在。"},
]

# 好感度是float，回傳的是dict
def get_affection_level(aff: float) -> dict:
    # 把current先設為最低的階層(陌生)
    current = AFFECTION_LEVELS[0]
    # 依序檢查好感度階層，當好感度大於該階層的門檻時，就把current改成該階層
    for lvl in AFFECTION_LEVELS:
        if aff >= lvl["threshold"]:
            current = lvl
    return current


# ════════════════════════════════════════════════
#  日常劇情字串
# ════════════════════════════════════════════════

# 根據不同好感度階層設定對應的日常劇情
DAILY_CHATS = {
    "陌生": [
        ["今天有讀書嗎？", "……保持習慣很重要。"],
        ["嗯。", "加油。"],
    ],
    "禮貌": [
        ["你今天讀了多久？", "不管多少，有讀就好。繼續努力。"],
        ["讀書辛苦了。", "記得休息。"],
        ["今天狀況還好嗎？", "學習要循序漸進，不用太急。"],
    ],
    "親近": [
        ["今天讀得怎麼樣？", "有沒有遇到什麼難題？"],
        ["你最近讀書的進度感覺不錯。", "繼續保持，我有在看的。"],
        ["……其實我也想多讀一點書。", "你能持續下去的話，讓我有點佩服你。"],
    ],
    "信任": [
        ["你知道嗎，我以前也有段時間完全讀不下去……", "後來是每天只逼自己讀十分鐘。", "你比我當時強多了。"],
        ["今天讀書了嗎。", "……我知道你會的。", "不知道為什麼，就是覺得你不會放棄。"],
        ["你有沒有想過，為什麼要這麼努力？", "我最近一直在想這個問題。", "……不用現在回答我，只是想說說。"],
    ],
    "親密": [
        ["今天也讀書了。", "……你真的很讓我安心。", "不知道為什麼這樣說，但就是這個感覺。"],
        ["嗯，我一直都有注意到你在努力。", "你知道嗎，每次看到你認真的樣子……", "算了，說出來會很奇怪。"],
        ["今天可以多讀一下嗎。", "我想多陪你一會兒。", "……這樣說很奇怪嗎？"],
        ["萬一哪天我不來這間自習室了？……那我就去你的系館找你。我已經習慣對面坐的人是你了，換成別人……我不習慣。"],
        ["我在便利商店看到這個軟糖，記得你上次多看了兩眼……就順手買了。給你吃。"],
        ["……可以牽嗎？對不起，我沒什麼經驗，但我現在……手心都是汗，因為太緊張了。"],
        ["今天自習室的冷氣開得有點強，我剛好有帶一件乾淨的備用針織背心，不介意的話可以披在腿上。"],
        ["我現在其實一個字都讀不下去。……因為你今天把頭髮紮起來了，我沒辦法不把視線移到你身上。"],
    ],
}


# ════════════════════════════════════════════════
#  主線劇情資料
#
#  結構說明：
#  STORY_PROLOGUE  : 共用序章（選路線前觸發，所有玩家都會看到）
#  STORIES         : dict，key 為 route 名稱，value 為該路線的章節 list
#
#  每個章節可以有：
#    text      : 純文字 list
#    text_fn   : lambda name -> list（含玩家名字時用）
#    branches  : 分歧支線 list（有此欄位時忽略 text/text_fn）
#      每條支線有：
#        id        : 支線唯一 id
#        subtitle  : 支線副標題（顯示在劇情畫面頂部）
#        condition : lambda state -> bool，None 表示 fallback
#        text / text_fn : 同上
#
#  excel_cell 規則：
#    共用序章用 H2；romance 路線從 H3 開始；horror 路線從 H20 開始
#    有分歧的章節，父章節記「是否看過」；
#    可在支線加 branch_cell 用來記錄走哪條支線（參考第三章範例）
# ════════════════════════════════════════════════

STORY_PROLOGUE = {
    "id": "story0",
    "title": "序章",
    "subtitle": "遊戲說明",
    "exp_threshold": 0,
    "requires": None,
    "excel_cell": "H2",
    "text": [
        "點擊左鍵繼續劇情",
        "歡迎使用 Study Timer。",
        "在這裡，每一次專注都將被記錄，而你的讀書時長，也會逐步推動故事發展。",
        "首先先簡單向您介紹遊戲機制",
        "遊戲內包含乙女向劇情與恐怖向劇情。",
        "讀書時可累積經驗值與好感度，經驗值將決定是否解鎖下一段劇情，而好感度則會影響角色互動與劇情分歧。",
        "每讀書 1 分鐘可獲得 1 點經驗值。",
        "當每日累積讀書時間超過 2 小時 後，後續獲得的經驗值將提升為 1.2 倍",
        "每連續讀書 15 分鐘可獲得 1 點好感度。",
        "當單次連續讀書時間超過 1 小時 後，在暫停前所獲得的好感度將以 2 倍計算",
        "好的~ 遊戲機制大致介紹完成啦",
        "接下來就請您開始讀書，並好好享受遊戲劇情吧！"
    ],
}

STORIES = {
    # ── 乙女向路線 ───────────────────────────────
    "romance": [
        {
            "id": "romance_ch1",
            "title": "第一章",
            "subtitle": "初次相遇",
            "exp_threshold": 60,
            "requires": None,
            "excel_cell": "H3",
            "text": [
                "午後的圖書館安靜得只剩下翻頁聲。",
                "你坐在靠窗的位置，桌上攤著還沒讀完的講義，筆記本邊角已經被反覆摩擦得微微捲起。",
                "陽光透過百葉窗切成一格一格的光影，落在紙面上，像是把時間也分割成細碎的片段。",
                "你揉了揉有些發酸的手腕，正準備繼續寫時，對面的位置忽然有人坐了下來。",
                "椅子輕輕發出一聲聲響。",
                "你下意識抬頭。",
                "那是一個看起來有點安靜過頭的男生。",
                "黑色微卷的頭髮有些蓬鬆，細框眼鏡後的視線短暫地與你交會，又很快地移開。",
                "他的動作有些僵硬，像是在努力讓自己看起來自然一點。",
                "「……抱、抱歉，這裡可以坐嗎？」",
                "他的聲音很輕，甚至帶著一點不易察覺的緊張。",
                "你點了點頭。",
                "他似乎鬆了一口氣，小心翼翼地把書放到桌上，動作俐落又整齊。",
                "你瞄了一眼封面——是一本厚重的醫學原文書，密密麻麻的英文讓人一看就覺得頭痛。",
                "他翻開書頁後，很快就進入狀態。",
                "原本略顯緊張的神情，在低頭閱讀時逐漸變得專注而穩定。",
                "筆尖落在紙上時，沒有一絲猶豫，像是早已在腦中整理好所有脈絡。",
                "你不自覺多看了兩眼。",
                "也許是察覺到你的視線，他的手微微停了一下，接著又繼續寫下去，但耳尖似乎悄悄地紅了。",
                "過了一會兒，他忽然把一張便條紙推到你桌邊。",
                "字跡乾淨整齊，帶著一點理性的距離感。",
                "「你剛剛那一題，公式寫錯了一個符號。」",
                "你愣了一下，低頭看向自己的筆記。",
                "——還真的錯了。",
                "當你再抬頭時，他已經重新低下頭，像是剛才那句話只是順手而為，沒有打算多做解釋。",
                "窗外的光慢慢移動。",
                "你忽然覺得，這個原本陌生的午後，好像悄悄多了一點不同的節奏。"
            ],
        },
        {
            "id": "romance_ch2",
            "title": "第二章",
            "subtitle": "不確定的相遇",
            "exp_threshold": 250,
            "requires": "romance_ch1",
            "excel_cell": "H4",
            "text_fn": lambda name: [
                "隔天，你還是來到了同一個位置。",
                "說不上為什麼，也許只是因為這裡的光線剛好",
                "也可能是因為——你隱約記得，昨天那個人坐在對面。",
                "你把書攤開，還沒完全進入狀態，對面的椅子就再次被拉開。",
                "熟悉的細微聲響。",
                "你抬頭。",
                "果然是他。",
                "他似乎也愣了一下，像是沒預料到你會在這裡，但很快地點了點頭，動作有些僵硬地坐下來。",
                "短暫的沉默在兩人之間擴散。",
                "你低頭寫了幾行筆記，又忍不住停下來。",
                "——總覺得，這樣對坐卻一句話都不說，有點奇怪。",
                "你還在猶豫的時候，對面的人先動了。",
                "他輕輕闔上書，指尖在書頁邊緣停了一下，像是在組織語言。",
                "「那個……」",
                "聲音比昨天還低了一點。",
                "他推了推眼鏡，視線沒有完全對上你。",
                "「昨天……我是不是有點失禮，只寫了紙條。」",
                "你愣了一下，搖頭。",
                "「不會，還好有你提醒。」",
                "他似乎鬆了一口氣，肩膀微微放鬆。",
                "沉默又出現了一瞬，但這次，他沒有讓它持續太久。",
                "「...我叫林霽安。」",
                "這句話說得比前面流暢一些，雖然尾音還是有點輕。",
                "「醫學系三年級。」",
                "說完之後，他像是完成了一件需要鼓起勇氣的事，目光短暫地落在你身上，又很快移開。",
                "接著，他遲疑了一下，小聲補上一句：",
                "「可以……知道你的名字嗎？」",
                f"「……我叫{name}。」",
                "他輕輕地重複了一次你的名字，語氣很認真，像是在確認發音。",
                "「……很好聽。」",
                "說完之後，他像是意識到這句話有點突然，耳尖微微紅了起來，立刻低下頭翻開書本，假裝專心。",
                "你忍不住笑了一下。",
                "氣氛比昨天鬆動了一點。",
                "之後的時間裡，你們沒有再多聊天。",
                "只是偶爾，你會感覺到對面傳來短暫的視線——",
                "在你皺眉的時候、停筆的時候，或是翻頁太快的時候。",
                "像是在確認什麼。",
                "也像是在默默記住。"
            ],
        },
        {
            # ── 分歧章節範例：好感度 ≥ 60 走親密支線，否則走普通支線 ──
            "id": "romance_ch3-1",
            "title": "第三章",
            "subtitle": "距離",           # 顯示於劇情進度頁；分歧後由 branch subtitle 覆蓋
            "exp_threshold": 500,
            "requires": "romance_ch2",
            "excel_cell": "H5",
            "branches": [
                {
                    "id": "romance_ch3_high",
                    "subtitle": "縮短的距離",
                    "condition": lambda state: state.affection >= 1,
                    "branch_cell": "H6",
                    "text": [
                        "那天，你比平常多讀了一段時間。",
                        "離開圖書館時，天色已經有點偏橘，校園裡的風也變得涼了一點。",
                        "你站在門口伸了個懶腰，還在猶豫要不要直接回去時，旁邊傳來熟悉的聲音。",
                        "「……你今天比較晚。」",
                        "你轉過頭。",
                        "林霽安站在不遠處，手裡拿著一罐還沒開的飲料，像是剛從販賣機那邊走過來。",
                        "「多讀了一點。」",
                        "他點了點頭，視線在你臉上停了一下。",
                        "「看得出來。」",
                        "你愣了一下，還沒來得及問，他就有點不自然地補了一句：",
                        "「就是……筆記寫很久的人，通常會那個表情。」",
                        "講完之後，他像是覺得自己解釋太多，微微別開視線。",
                        "短暫的沉默後，他把手上的飲料遞過來。",
                        "「無糖的茶。」",
                        "他語氣平靜地說。",
                        "你有點意外地接過。",
                        "冰涼的觸感從掌心傳上來。",
                        "「謝謝。」",
                        "「……嗯。」",
                        "他應了一聲，很輕。",
                        "你們一起走到旁邊的長椅坐下。",
                        "傍晚的風慢慢吹過來，帶走一點長時間讀書的疲憊。",
                        "你打開飲料喝了一口。",
                        "過了一會兒，他忽然開口：",
                        "「你今天，中間沒有滑手機。」",
                        "你轉頭看他。",
                        "「你怎麼知道？」",
                        "他愣了一下，像是沒想到你會直接問。",
                        "「……我剛好有看到幾次。」",
                        "他停頓了一下。",
                        "「你幾乎都在寫。」",
                        "這句話說得很慢，但很確定。",
                        "「這樣其實滿好的，長時間專注……效率會比較穩。」",
                        "語氣已經不像一開始那樣緊張，反而帶著一點他在熟悉領域時的沉著。",
                        "你忍不住笑了。",
                        "「你是在觀察我嗎？」",
                        "話一出口，他整個人明顯僵了一下。",
                        "「我、我沒有——」",
                        "說到一半，他停住，像是覺得否認也不太對。",
                        "耳尖慢慢紅起來。",
                        "「……只是剛好注意到。」",
                        "風從你們之間吹過。",
                        "這次的沉默，沒有尷尬。",
                        "反而有點剛剛好的距離。"
                    ],
                },
                {
                    "id": "romance_ch3_low",
                    "subtitle": "維持的距離",
                    "condition": None,      # fallback，不符合上面條件時走這裡
                    "text_fn": lambda name: [
                        "你離開圖書館的時候，天色還亮著。",
                        "讀書的節奏斷斷續續，時間過得有點快，但內容卻沒怎麼記住。",
                        "你站在門口，猶豫了一下，還是往外走。",
                        "沒走幾步，就看到熟悉的身影。",
                        "林霽安站在販賣機前，低頭看著按鈕，像是在思考要選哪一個。",
                        "你稍微放慢腳步。",
                        "他也注意到了你。",
                        f"「……{name}。」",
                        "他點了一下頭。",
                        "你回應了一聲。",
                        "短暫的沉默。",
                        "機器發出「咚」的一聲，他彎腰拿起掉下來的飲料。",
                        "你們之間隔著一點距離。",
                        "不像之前在同一張桌子那樣自然。",
                        "「你今天……比較早。」",
                        "他開口。",
                        "「嗯，有點讀不太下去。」",
                        "你坦白說。",
                        "他愣了一下。",
                        "手上的動作停住一瞬。",
                        "「……這樣啊。」",
                        "語氣沒有評價，只是很輕地接住這句話。",
                        "他沒有再多問。",
                        "只是把飲料拿好，稍微側身讓出位置。",
                        "「那我先走了。」",
                        "你點點頭。",
                        "「嗯。」",
                        "你們沒有一起走。",
                        "他往圖書館的方向，你往校門口。",
                        "擦肩而過的時候，你忽然覺得，好像有什麼本來可以發生的對話，被錯過了。",
                        "但你也說不上來，是什麼。"
                    ],
                },
            ],
        },
        {
            "id": "romance_ch4",
            "title": "第四章",
            "subtitle": "順其自然",
            "exp_threshold": 700,
            "requires": "romance_ch3-1",
            "excel_cell": "H7",
            "text_fn":lambda name:[
                "你開始不再特別思考要不要去那個地方。",
                "下課之後，收好東西，腳步自然地往同一棟系館走去。",
                "自習室在三樓，燈光比圖書館暖一些，人也少一點。",
                "你推開門的時候，下意識看向靠窗的第三排。",
                "桌上已經放了一本書。",
                "旁邊留出一個剛好的空位。",
                "你走過去，還沒坐下，就聽見身後熟悉的聲音。",
                f"「...{name}，你來了。」",
                "你回頭。",
                "林霽安站在門口，手裡拿著水。",
                "你看了看桌上的書。",
                "「你幫我留位置？」",
                "他愣了一下。",
                "「我只是……剛好坐這裡。」",
                "語氣依舊有點不自然。",
                "你坐了下來。",
                "「謝啦。」",
                "他輕輕點頭，在你對面坐下。",
                "時間在熟悉的節奏中流動。",
                "沒有特別的對話，但你們的動作卻慢慢變得一致。",
                "翻頁、停筆、休息——",
                "像是不用說出口的默契。",
                "過了一段時間，窗外的光線變暗了。",
                "你原本以為只是天色晚了。",
                "直到雨聲落下。",
                "一開始只是零碎的幾滴。",
                "很快地，變成連續的聲響。",
                "啪嗒、啪嗒——",
                "打在窗戶上，也打在你的注意力上。",
                "你抬頭看了一眼。",
                "「下雨了。」",
                "「……嗯。」",
                "他也停下筆，看向窗外。",
                "雨勢不小。",
                "你皺了一下眉。",
                "「我沒帶傘。」",
                "話一出口，你才意識到自己說得太自然了。",
                "但他沒有覺得奇怪。",
                "只是停頓了一下。",
                "「……我也是。」",
                "短暫的安靜。",
                "你們對看了一眼，又同時移開視線。",
                "—",
                "雨沒有變小的意思。",
                "你們還是把該讀的部分收尾，才一起離開自習室。",
                "走到一樓時，外面的雨聲更清楚了。",
                "屋簷外是一整片密集的雨幕。",
                "你們停在門口。",
                "誰都沒有先走出去。",
                "「看起來……要等一下。」",
                "你說。",
                "「……嗯。」",
                "他應了一聲。",
                "你們站在同一段屋簷下。",
                "距離不遠，也不算太近。",
                "雨聲把世界隔開。",
                "路上的人變少，聲音被沖淡，只剩下規律的落水聲。",
                "風從側邊吹進來，帶著一點雨水的涼意。",
                "你下意識往裡站了一點。",
                "他注意到了這個動作。身體微微動了一下，像是想讓出更多空間。",
                "但又沒有真的移動太多。",
                "像是在拿捏一個剛剛好的距離。",
                "雨還在下。",
                "時間被拉長。",
                "你們沒有再多說話。",
                "卻也沒有不自在。",
                "只是安靜地站在同一個地方，等雨停。",
                "某個瞬間，你忽然覺得——",
                "這樣的沉默，好像已經和一開始不一樣了。"
            ]
        },
        {
            "id": "romance_ch5",
            "title": "第五章",
            "subtitle": "細雨之間",            # 進度頁顯示用，分歧後由 branch subtitle 覆蓋
            "exp_threshold": 1200,
            "requires": "romance_ch4",
            "excel_cell": "H8",
            "branches": [
                {
                    "id": "romance_ch5_high",
                    "subtitle": "雨聲漸歇",
                    "condition": lambda state: state.affection >= 65,
                    "branch_cell": "H9",
                    "text": [
                        "雨勢終於慢慢變小。",
                        "從一整片模糊的水幕，變成斷斷續續的細雨。",
                        "屋簷外的地面還濕著，空氣裡帶著一點涼意。",
                        "你往外看了一眼。",
                        "「好像可以走了。」",
                        "林霽安沒有立刻回應。",
                        "他也看著外面，像是在判斷什麼。",
                        "幾秒後，他開口。",
                        "「……你往哪個方向？」",
                        "你報上了大概的路線。",
                        "他點了點頭。",
                        "「有一段……順路。」",
                        "他停頓了一下。",
                        "像是在組織語言。",
                        "「如果你不介意的話，可以……一起走到那邊。」",
                        "語氣還是很輕。",
                        "但比以前少了一點猶豫。",
                        "你愣了一下，點頭。",
                        "「好啊。」",
                        "你們一起走進雨裡。",
                        "雨已經不大，但還是會落在肩上、手臂上。",
                        "你下意識縮了一下。",
                        "走在你旁邊的人，似乎注意到了。",
                        "他的步伐微微調整，站在比較靠外側的位置。",
                        "風從那一側吹過來，大部分的雨也被擋住了一點。",
                        "你沒有說破。",
                        "只是默默跟著他的步調。",
                        "路不算長。",
                        "但也沒有短到讓人覺得尷尬。",
                        "走到岔路口時，他停了下來。",
                        "「我到這裡。」他說。",
                        "你點點頭。",
                        "「今天……謝謝。」",
                        "這句話出口時，你自己都不太確定是在謝什麼。",
                        "也許是那段一起讀書的時間。",
                        "也許是剛剛那段路。",
                        "他愣了一下。",
                        "「……沒什麼。」",
                        "他輕聲說。",
                        "停了一秒，又補上一句：",
                        "「你今天，讀得很好。」",
                        "這句話說得很自然。",
                        "不像以前那樣需要鼓起勇氣。",
                        "你笑了。",
                        "「被你觀察到了？」",
                        "他微微一僵。",
                        "然後，小幅度地點了點頭。",
                        "這次，沒有否認。",
                        "雨還在下。",
                        "但你忽然覺得，好像沒有那麼冷了。"
                    ],
                },
                {
                    "id": "romance_ch5_low",
                    "subtitle": "未曾落定",
                    "condition": None,    # fallback
                    "branch_cell": "H10",
                    "text": [
                        "雨勢慢慢變小。",
                        "屋簷外從模糊的水幕，變成細碎的雨絲。",
                        "你往外看了一眼。",
                        "「應該可以走了。」",
                        "林霽安點了點頭。",
                        "「……嗯。」",
                        "短暫的停頓。",
                        "像是在等誰再說一句話。",
                        "但沒有人開口。",
                        "你往前走了一步。",
                        "他也跟著走出屋簷。",
                        "雨還在落。",
                        "不大，但足以讓人不太舒服。",
                        "你們走在同一條路上。",
                        "距離不遠。",
                        "卻沒有像剛才那樣站在同一個空間裡的感覺。",
                        "你想說點什麼。",
                        "卻又覺得，好像沒有一個自然的開頭。",
                        "他也沒有開口。",
                        "只是維持著一樣的步伐。",
                        "到了分岔路口，你停下來。",
                        "「我往這邊。」你說。",
                        "他點點頭。",
                        "「……好。」",
                        "又是一小段安靜。",
                        "你抬頭看了他一眼。",
                        "「那，明天見？」",
                        "這句話帶著一點試探。",
                        "他愣了一下。",
                        "「……嗯。」",
                        "回應不慢。",
                        "但也沒有延伸。",
                        "你點點頭，轉身離開。",
                        "雨落在肩上，有點冰。",
                        "你走了幾步，忍不住回頭看了一眼。",
                        "他還站在原地。",
                        "像是也在猶豫什麼。",
                        "但最後，還是轉身往另一個方向走去。",
                        "你忽然意識到——",
                        "你們之間，明明已經不算陌生。",
                        "卻還沒有誰，真正往前走一步。"
                    ],
                },
            ],
        },
    ],
    

    # ── 恐怖向路線 ───────────────────────────────
    "horror": [
        {
            "id": "horror_ch1",
            "title": "第一章",
            "subtitle": "閉館後",
            "exp_threshold": 100,
            "requires": None,
            "excel_cell": "H20",
            "text_fn": lambda name: [
                "不知道從什麼時候開始，圖書館裡的燈少了一盞。",
                "我抬起頭，走廊的盡頭是黑的。",
                "剛才……那裡有人嗎？",
                f"（{name}，你聽到了嗎？）",
                "我搖搖頭，繼續低頭看書。",
                "空調嗡嗡作響。",
                "書頁自己翻了一頁。",
                "我沒有碰它。",
            ],
        },
        {
            "id": "horror_ch2",
            "title": "第二章",
            "subtitle": "不該翻開的那頁",
            "exp_threshold": 250,
            "requires": "horror_ch1",
            "excel_cell": "H11",
            "text_fn": lambda name: [
                "那本書還在桌上。",
                "我昨天明明把它放回書架了。",
                "封面上沒有書名，只有一個指紋——",
                "和我的一模一樣。",
                f"（{name}……）",
                "有聲音在叫我的名字。",
                "我沒有告訴任何人我叫什麼。",
            ],
        },
        {
            # ── 恐怖向分歧：好感度代表「與異常的距離」，高好感反而走更深的支線 ──
            "id": "horror_ch3",
            "title": "第三章",
            "subtitle": "???",
            "exp_threshold": 500,
            "requires": "horror_ch2",
            "excel_cell": "H12",
            "branches": [
                {
                    "id": "horror_ch3_deep",
                    "subtitle": "你翻開了它",
                    "condition": lambda state: state.affection >= 40,
                    "branch_cell": "H13",
                    "text_fn": lambda name: [
                        "我翻開了那本書。",
                        "裡面寫的是我的名字。",
                        "每一頁都是。",
                        f"「{name}，你終於來了。」",
                        "我不知道那個聲音從哪裡來。",
                        "但我知道——",
                        "它一直在等我。",
                    ],
                },
                {
                    "id": "horror_ch3_escape",
                    "subtitle": "你放下了它",
                    "condition": None,
                    "text_fn": lambda name: [
                        "我沒有翻開那本書。",
                        "我把它推到桌子邊緣，站起來，走向出口。",
                        "門沒有鎖。",
                        "外面的空氣是正常的。",
                        f"（{name}，你做了正確的選擇。）",
                        "但我不知道是誰說的。",
                        "我回頭看了一眼。",
                        "桌上，書不見了。",
                    ],
                },
            ],
        },
        # 繼續新增恐怖向章節請照上方格式，excel_cell 接著用 H14、H15...
    ],
}


# ════════════════════════════════════════════════
#  取得當前路線的章節 list
# ════════════════════════════════════════════════

def get_active_stories(state) -> list:
    """回傳當前故事線的章節 list；尚未選擇時回傳空 list。"""
    # 從state抓出玩家選擇的路線，然後到STORIES找出這個路線的所有章節
    # 回傳[]是為了避免開局玩家還沒選擇路線時程式崩潰
    return STORIES.get(state.story_route, [])

def get_all_story_cells() -> list:
    """回傳所有章節（含所有路線）用到的 excel_cell，供 init_excel 初始化用。"""
    # 儲存是否有看過序章的excel儲存格
    cells = [STORY_PROLOGUE["excel_cell"]]
    # 跑遍STORIES中的所有路線
    for route_stories in STORIES.values():
        # 把每個路線儲存劇情紀錄的儲存格找出來
        for s in route_stories:
            cells.append(s["excel_cell"])
            # 如果有支線劇情，就把據路是否有看過支線劇情的儲存格也找出來
            for b in s.get("branches", []):
                if b.get("branch_cell"):
                    cells.append(b["branch_cell"])
    return cells


# ════════════════════════════════════════════════
#  分歧劇情選擇
# ════════════════════════════════════════════════

def pick_branch(branches: list, state) -> dict:
    """從上往下找第一個 condition 為 True 的支線；都不符合則用 fallback（condition=None）。"""
    fallback = None
    for branch in branches:
        # 先找條件為None的分支當保底
        if branch["condition"] is None:
            fallback = branch
        # branch["condition"]是在STORIES中寫好的函數(確認好感度是否達標)
        # 把state丟進函數中檢查是否有達到解鎖該分支的好感度門檻
        # 有的話就回傳該分支
        elif branch["condition"](state):
            return branch
    # 沒有達到其他分支的門檻時就回傳保底劇情
    # branches[-1]是為了避免前面分支忘記設保底劇情導致程式出錯
    # 之後檢查分支設定無誤時可以把branches[-1]去掉
    return fallback or branches[-1]


# ════════════════════════════════════════════════
#  EXP 計算
# ════════════════════════════════════════════════
# 設定每日可獲取的經驗值上限
# 經驗值決定劇情是否解鎖，好感度決定解鎖的分支劇情
# 如果在好感度達標之前就解鎖劇情的話，就只能看到比較差的劇情
# 在這個機制下，經驗值達到獲取上限後仍可以獲取好感度
# 算是給想解鎖更好的分支劇情，但好感度沒有達標的玩家一條出路
# 找一天多讀一點書，讀到經驗值獲取上限後就可以在不增加經驗值的情況下累積好感度
DAILY_EXP_CAP = 200

def calc_exp(actual_secs: float, daily_exp_so_far: float = 0) -> float:
    # 單次讀書時間
    actual_min = actual_secs / 60
    # 經驗值加成門檻
    BONUS_THRESHOLD = 120
    # 如果今日讀書的總時長已超過門檻，直接把這次讀的分鐘數*1.2
    # daily_exp_so_far=120就代表剛好讀了2小時
    # 這裡用daily_exp_so_far是因為儲存遊戲紀錄的excel表中沒有每日讀書時間的紀錄
    if daily_exp_so_far >= BONUS_THRESHOLD:
        raw_exp = actual_min * 1.2
    # 如果讀書時長剛好跨越了2小時的門檻
    elif daily_exp_so_far + actual_min > BONUS_THRESHOLD:
        # 未達兩小時的分鐘數正常算
        before = BONUS_THRESHOLD - daily_exp_so_far
        # 超過兩小時的*1.2倍
        after  = actual_min - before
        # 把兩個結果相加
        raw_exp = before + after * 1.2
    else:
        # 未達兩小時=>正常算
        raw_exp = actual_min
    # DAILY_EXP_CAP - daily_exp_so_far代表今天還可獲得的經驗值
    # 其實這裡好像不需要用max...?因為daily_exp_so_far最高就是200
    # 但我不敢動XD
    remaining_cap = max(0, DAILY_EXP_CAP - daily_exp_so_far)
    # 從raw_exp和remaining_cap回傳最小的數值
    return min(raw_exp, remaining_cap)


# ════════════════════════════════════════════════
#  好感度計算
# ════════════════════════════════════════════════
# 每天沒上線扣的好感度
AFF_DECAY_PER_DAY  = 5
AFF_MIN, AFF_MAX   = 0, 100

def calc_affection_gain(actual_secs: float) -> float:
    """
    每 15 分鐘 +1 好感度。
    連續讀超過 60 分鐘後，剩餘部分的好感度 × 2。
    """
    actual_min = actual_secs / 60
    # 連續讀超過 1 小時後開始加成
    BONUS_THRESHOLD = 60 
    if actual_min <= BONUS_THRESHOLD:
        # 沒超過 1 小時，直接算
        aff = actual_min / 15
    else:
        # 前 60 分鐘正常算，超過的部分 × 2
        before = BONUS_THRESHOLD / 15
        after  = (actual_min - BONUS_THRESHOLD) / 15 * 2
        aff = before + after

    return aff

# 好感度減少計算
def calc_affection_decay(days_absent):
    # 最多只扣連續4天沒上線的好感度
    return AFF_DECAY_PER_DAY * min(days_absent, 4)


# ════════════════════════════════════════════════
#  AppState
# ════════════════════════════════════════════════

# 定義遊戲資料的class
@dataclass
class AppState:
    start_time: float = None
    elapsed_time: float = 0
    total_time: float = 0
    average_time: float = 0
    running: bool = False
    session_count: int = 0
    user_name: str = ""
    game_data: str = "study_time.xlsx"
    story_flags: dict = field(default_factory=dict)
    total_exp: float = 0
    daily_exp: float = 0
    affection: float = 0
    daily_aff_gained: float = 0
    last_study_date: str = ""
    today_chat_shown: bool = False
    story_route: str = ""           # "romance" | "horror" | ""


# ════════════════════════════════════════════════
#  時間工具
# ════════════════════════════════════════════════

# 把總秒數轉換成 HH:MM:SS 格式的字串
def fmt(secs):
    secs = max(0, int(secs))
    return f"{secs//3600:02}:{(secs%3600)//60:02}:{secs%60:02}"

# 把總秒數換算成 MM:SS 格式的字串(適合一小時以內)
def fmt_min(secs):
    secs = max(0, int(secs))
    return f"{secs//60:02}:{secs%60:02}"

# 把時間字串換成秒數
def time_str_to_sec(ts):
    h, m, s = map(int, ts.split(":"))
    return h * 3600 + m * 60 + s

# 把時間字串換成分鐘數
def time_str_to_min(ts):
    h, m, s = map(int, ts.split(":"))
    return h * 60 + m + s / 60

# 取得今日時間的字串
def today_str():
    return datetime.date.today().strftime("%Y-%m-%d")


# ════════════════════════════════════════════════
#  Excel 存取函式
# ════════════════════════════════════════════════

# 初次登入時建立保存遊戲資料的excel表
def init_excel(game_data):
    if not os.path.exists(game_data):
        wb = Workbook()
        ws = wb.active
        ws.title = "紀錄"
        ws["A1"] = "總時間";        ws["B1"] = "單次讀書時間"
        ws["C1"] = "計時次數";      ws["C3"] = "平均讀書時間"
        ws["D1"] = "累積EXP";       ws["D2"] = 0
        ws["E1"] = "今日EXP";       ws["E2"] = 0
        ws["F1"] = "好感度";        ws["F2"] = 0
        ws["G1"] = "今日好感度增量"; ws["G2"] = 0
        ws["H1"] = "劇情旗標"
        ws["I1"] = "使用者名稱"
        ws["J1"] = "上次讀書日期";  ws["J2"] = ""
        ws["K1"] = "今日日常已顯示"; ws["K2"] = False
        ws["L1"] = "故事線";        ws["L2"] = ""
        for cell in get_all_story_cells():
            ws[cell] = False
        wb.save(game_data)

# 讀檔/存檔初始化系統
def load_excel_data(game_data):
    wb = load_workbook(game_data)
    ws = wb["紀錄"]

    # 收集所有可能的章節 id（含所有路線）供 story_flags 初始化
    all_story_ids = [STORY_PROLOGUE["id"]]
    # 找出STORIES中的所有路線
    for route_stories in STORIES.values():
        # 找出路線中的所有劇情
        for s in route_stories:
            # 把每個劇情的id找出來
            all_story_ids.append(s["id"])
            # 有分支的話就把分支的劇情id也找出來
            for b in s.get("branches", []):
                all_story_ids.append(b["id"])

    # 遊戲初始化設定
    result = {
        "total_time": 0, "session_count": 0, "average_time": 0,
        "user_name": "", "story_route": "",
        # 把劇情id轉成以下結構："story_flags": {"id_01": False, "id_02":False, ...}
        "story_flags": {sid: False for sid in all_story_ids},
        "total_exp": 0, "daily_exp": 0, "affection": 0,
        "daily_aff_gained": 0, "last_study_date": "", "today_chat_shown": False,
    }

    # 如果讀書總時間和總計時次數的儲存格已有數值，就載入total_time和session_count的紀錄
    if ws["A2"].value is not None and ws["C2"].value is not None:
        result["total_time"]    = time_str_to_sec(ws["A2"].value)
        result["session_count"] = int(ws["C2"].value)
        # 如果平均讀書時間已有數值，就載入average_time的紀錄
        if ws["C4"].value is not None:
            result["average_time"] = time_str_to_sec(ws["C4"].value)

    # 讀取所有章節旗標（含序章、所有路線章節、支線旗標）
    flag_map = {STORY_PROLOGUE["excel_cell"]: STORY_PROLOGUE["id"]}
    for route_stories in STORIES.values():
        for s in route_stories:
            # 找出每一段劇情觀看紀錄儲存的位置
            # flag_map的結果會長得像這樣：{H3:"id_01", H4:"id_02", ...}
            flag_map[s["excel_cell"]] = s["id"]
            for b in s.get("branches", []):
                if b.get("branch_cell"):
                    # 找出分支劇情觀看紀錄儲存的位置
                    flag_map[b["branch_cell"]] = b["id"]
    for cell, sid in flag_map.items():
        # 讀取excel，找到儲存劇情紀錄的格子中的數值
        val = ws[cell].value
        # 如果格子裡有數值代表玩家看過該段劇情，result["story_flags"][sid]改成True，反之改成False
        result["story_flags"][sid] = bool(val) if val is not None else False

    # 讀取使用者名稱
    if ws["I2"].value is not None:
        result["user_name"] = ws["I2"].value

    # 讀取其他資料
    # 先列出每個資料存放的位置，以及資料的type
    # 資料型態、名稱和儲存位置基本上不會更動，所以用tuple包起來
    for cell, key, cast in [
        ("D2", "total_exp",        float),
        ("E2", "daily_exp",        float),
        ("F2", "affection",        float),
        ("G2", "daily_aff_gained", float),
        ("J2", "last_study_date",  str),
        ("K2", "today_chat_shown", bool),
        ("L2", "story_route",      str),
    ]:
        # 讀取資料並傳入result中
        val = ws[cell].value
        if val is not None:
            result[key] = cast(val) if cast != bool else bool(val)

    return result


# 儲存/更新遊戲紀錄(每次計時結束後使用)
def save_session(game_data, total_time, elapsed_time, session_count,
                 average_time, total_exp, daily_exp, affection,
                 daily_aff_gained, last_study_date):
    wb = load_workbook(game_data)
    ws = wb["紀錄"]
    # row的初始值設定為2(第一筆單次讀書紀錄在B2)
    row = 2
    # 第B欄是用來記錄單次讀書時間的，每次都要往下一格紀錄
    # 從B2開始往下找，如果欄位已有紀錄就往下一格，直到找到沒有紀錄的欄位為止
    while ws[f"B{row}"].value is not None:
        row += 1
    ws["A2"] = fmt(total_time);          ws[f"B{row}"] = fmt(elapsed_time)
    ws["C2"] = session_count;            ws["C4"] = fmt(average_time)
    ws["D2"] = round(total_exp, 2);      ws["E2"] = round(daily_exp, 2)
    ws["F2"] = round(affection, 2);      ws["G2"] = round(daily_aff_gained, 2)
    ws["J2"] = last_study_date
    wb.save(game_data)


# 讀完劇情後記錄的函式
def save_story_flag(game_data, cell, value):
    wb = load_workbook(game_data)
    ws = wb["紀錄"]
    ws[cell] = value
    wb.save(game_data)


# 儲存使用者名稱的函式
def save_user_name(game_data, name):
    wb = load_workbook(game_data)
    ws = wb["紀錄"]
    ws["I2"] = name
    wb.save(game_data)


# 儲存劇情路線的函式
def save_story_route(game_data, route):
    wb = load_workbook(game_data)
    ws = wb["紀錄"]
    ws["L2"] = route
    wb.save(game_data)


# 儲存是否觸發每式劇情的函式
def save_today_chat_shown(game_data):
    wb = load_workbook(game_data)
    ws = wb["紀錄"]
    ws["K2"] = True
    wb.save(game_data)


# 載入之前每次讀書紀錄
def load_session_history(game_data):
    wb = load_workbook(game_data)
    ws = wb["紀錄"]
    # 先建一個空list
    sessions = []
    # 從第2列開始記錄(單次讀書時間從B2開始記錄)
    row = 2
    # 只要該儲存格有數值，就append到session中
    while ws[f"B{row}"].value is not None:
        sessions.append(ws[f"B{row}"].value)
        # apeend後把row+1去找下一個欄位
        row += 1
    return sessions


# ════════════════════════════════════════════════
#  好感度衰減
# ════════════════════════════════════════════════

def apply_daily_decay(state):
    # 取得今天的日期字串
    today = today_str()
    # 更新每日經驗值、好感度以及日常劇情
    if state.last_study_date != today:
        state.daily_exp = 0
        state.daily_aff_gained = 0
        state.today_chat_shown = False
    # 如果沒有last_study_date(新帳號)或是使用者今天已經登入了(last_study_date == today)就不需要扣好感度
    if not state.last_study_date or state.last_study_date == today:
        return 0.0
    # 把上次上線的日期字串轉回成可以計算的日期物件
    last = datetime.date.fromisoformat(state.last_study_date)
    # 計算缺席的天數
    days_absent = (datetime.date.today() - last).days - 1
    # 沒有缺席就不扣好感度
    if days_absent <= 0:
        return 0.0
    # 有缺席就使用calc_affection_decay計算要扣的好感度
    decay = calc_affection_decay(days_absent)
    # 扣除好感度
    state.affection = max(AFF_MIN, state.affection - decay)
    return decay


# ════════════════════════════════════════════════
#  劇情觸發判定
# ════════════════════════════════════════════════

def check_story_triggers(state, on_story=None):
    """依序檢查當前路線的章節，遇到第一個應觸發的就執行並返回。"""
    # 呼叫 get_active_stories(state) 拿到目前路線的所有章節
    for s in get_active_stories(state):
        sid = s["id"]
        # 已經觸發過的劇情就跳過
        if state.story_flags.get(sid):
            continue
        # 取得觸發該劇情的requires(要先觸發前面的章節)
        req = s.get("requires")
        # 沒有達成requires的劇情也跳過
        if req and not state.story_flags.get(req):
            continue
        # 如果經驗值大於門檻就觸發劇情
        if state.total_exp >= s["exp_threshold"]:
            if on_story:
                # 如果該段劇情有支線就用pick_branch判定觸發的支線
                if "branches" in s:
                    chosen = pick_branch(s["branches"], state)
                    # 把支線資料合併進父章節 dict，show_story 只看 title/subtitle/text/text_fn
                    # 把「原本主章節的資料（`**s`）」和「被選中的支線劇情資料（`**chosen`）」重疊融合在一起。
                    # 有點難解釋，但你們可以去問AI好朋友
                    merged = {
                        **s,
                        **chosen,
                        "parent_id": s["id"],        # ← 保留父章節 id
                        "branch_id": chosen["id"],   # ← 明確記支線 id
                    }
                    on_story(merged)
                else:
                    # 沒有支線就單純觸發劇情
                    on_story(s)
            return True       # ← 有觸發就回傳 True
    return False          # ← 沒有可觸發的回傳 False


# ════════════════════════════════════════════════
#  計時器邏輯
# ════════════════════════════════════════════════

# 開始計時的函式
def start_timer(state):
    # 如果目前不處於計時狀況，就開始計時
    if not state.running:
        # time.time()會回傳現在的時間戳記
        # elapsed_time 是過去累積專注的時間
        # 兩者相減用來處理暫停後繼續計時的讀書時數
        state.start_time = time.time() - state.elapsed_time
        state.running = True


# 暫停計時的函數
def do_pause(state, callbacks):
    # 如果不處於計時狀態案暫停就不用做任何事
    if not state.running:
        return
    # 如果原本計時器有再跑，暫停後就把計時器關掉(state.running = False)
    state.running = False
    # 把這一次專注累積的秒數拿出來計算經驗值
    actual_secs = state.elapsed_time
    exp_gained  = calc_exp(actual_secs, state.daily_exp)
    state.total_exp += exp_gained;  state.daily_exp += exp_gained
    # 計算增加的好感度
    aff_gained  = calc_affection_gain(actual_secs)
    state.affection = min(AFF_MAX, state.affection + aff_gained)
    state.daily_aff_gained += aff_gained
    # 更新總計時時間
    state.total_time    += actual_secs
    # 更新計時次數
    state.session_count += 1
    # 更新平均讀書時間
    state.average_time   = state.total_time / state.session_count
    # 更新上次讀書的日期
    state.last_study_date = today_str()
    # 儲存紀錄
    save_session(state.game_data, state.total_time, actual_secs,
                 state.session_count, state.average_time,
                 state.total_exp, state.daily_exp,
                 state.affection, state.daily_aff_gained, state.last_study_date)
    # 把 elapsed_time 的值設回零(下一次計時用)
    state.elapsed_time = 0
    # 彈出視窗告訴玩家剛剛計時獲得的經驗值和好感度
    if callbacks.get("on_pause"):
        callbacks["on_pause"](exp_gained, aff_gained)
    # 確認是否有劇情被觸發
    check_story_triggers(state, callbacks.get("on_story"))


# ════════════════════════════════════════════════
#  劇情顯示
# ════════════════════════════════════════════════

# 取得劇情文字
def get_story_text(story_def, user_name):
    # 如果劇情需要顯示使用者名稱，就把user_name傳進去
    if "text_fn" in story_def:
        return story_def["text_fn"](user_name)
    return story_def["text"]


# 顯示劇情
def show_story(story_def, user_name, story_frame, prev_frame,
               route="romance", on_no_finish=None, on_finished=None):
    """
    顯示劇情介面。
    route 參數決定背景色與文字色（乙女向 vs 恐怖向）。
    """
    # 取得玩家選擇的路線，再去ROUTE_CONFIG抓對應的介面設定(預設是乙女向)
    cfg = ROUTE_CONFIG.get(route, ROUTE_CONFIG["romance"])
    # 設定背景顏色
    story_bg = cfg["story_bg"]
    # 設定文字顏色
    story_fg = cfg["story_fg"]

    # 隱藏上一個介面
    prev_frame.pack_forget()
    # 清空上一個劇情頁面的文字和按鈕
    for w in story_frame.winfo_children():
        w.destroy()
    # 把劇情畫面的背景色換成剛剛決定的主題色
    story_frame.configure(bg=story_bg)
    # 把頁面放大到整個視窗
    story_frame.pack(fill="both", expand=True)

    # 取得劇情的文字
    text_list = get_story_text(story_def, user_name)
    current = {"index": 0}

    # 在劇情頁面上顯示標題與副標題
    tk.Label(story_frame,
             # text用來輸入要顯示的文字
             text=f"{story_def['title']}　{story_def['subtitle']}",
             # font 輸入文字設定
             font=("Georgia", 13, "italic"),
             # 設定背景和文字顏色(文字顏色根據選擇的路線改變)
             bg=story_bg, fg=C["pink_text"] if route == "romance" else "#9B7FA6"
             ).place(x=20, y=18)
    # 拉出一條分割線，把標題和底下的對話框隔開。
    tk.Frame(story_frame, height=1, bg=C["border"]).place(x=20, y=44, width=360)

    # 顯示第一段劇情文字(justify="left"代表靠左對齊)
    # wraplength=330代表文字超過330像素就自動換行(避免文字超出視窗)
    # padx、pady排版用
    text_label = tk.Label(story_frame, text=text_list[0],
                          wraplength=330, font=("Georgia", 15),
                          justify="left", bg=story_bg, fg=story_fg,
                          padx=30, pady=20)
    text_label.place(x=0, y=60, width=400, height=480)

    # 進度標籤，顯示還有多少句劇情
    prog_label = tk.Label(story_frame, text=f"1 / {len(text_list)}",
                          font=("Helvetica Neue", 11),
                          bg=story_bg, fg=C["gray"])
    prog_label.place(x=0, y=560, width=400)

    # 提示標籤，在視窗下方顯示「點擊畫面繼續 →」的提示
    hint_color = C["pink_light"] if route == "romance" else "#9B7FA6"
    hint = tk.Label(story_frame, text="點擊畫面繼續 →",
                    font=("Helvetica Neue", 11),
                    bg=story_bg, fg=hint_color)
    hint.place(x=0, y=585, width=400)

    # 顯示下一句劇情
    def next_text(event=None):
        current["index"] += 1
        # 如果劇情還沒播完，就換下一句劇情的文字並更新下方的進度標籤
        if current["index"] < len(text_list):
            text_label.config(text=text_list[current["index"]])
            prog_label.config(text=f"{current['index']+1} / {len(text_list)}")
   
        else:
            # 如果播完劇情就把劇情頁面隱藏
            if on_finished:
                on_finished()
            story_frame.pack_forget()

    # 偵測背景、對話框、進度條、提示文字
    # 只要玩家點擊這些地方就執行next_text
    for w in [story_frame, text_label, prog_label, hint]:
        w.bind("<Button-1>", next_text)


# ════════════════════════════════════════════════
#  UI 工具
# ════════════════════════════════════════════════

# 清除頁面上的文字和按鈕
def clear_frame(frame):
    for w in frame.winfo_children():
        w.destroy()


# 在視窗上畫一條水平線
def make_sep(parent, bg_color=None):
    return tk.Frame(parent, height=1, bg=bg_color or C["border"])


# ════════════════════════════════════════════════
#  主程式
# ════════════════════════════════════════════════

def main():
    # 把state定義成之前設好的class
    state = AppState()
    

    #創建/載入遊戲資料
    init_excel(state.game_data)
    saved = load_excel_data(state.game_data)
    # 把遊戲資料輸入到state裡
    state.total_time       = saved["total_time"]
    state.session_count    = saved["session_count"]
    state.average_time     = saved["average_time"]
    state.user_name        = saved["user_name"]
    state.story_flags      = saved["story_flags"]
    state.total_exp        = saved["total_exp"]
    state.daily_exp        = saved["daily_exp"]
    state.affection        = saved["affection"]
    state.daily_aff_gained = saved["daily_aff_gained"]
    state.last_study_date  = saved["last_study_date"]
    state.today_chat_shown = saved["today_chat_shown"]
    state.story_route      = saved["story_route"]

    # 計算衰減的好感度
    decay_amount = apply_daily_decay(state)

    # 創建基礎視窗
    root = tk.Tk()
    # 把視窗標題設為讀書計時器
    root.title("讀書計時器")
    # 設定視窗的初始寬度與高度
    root.geometry(f"{W}x{H}")
    # 把整個視窗的背景色漆成預設的底色
    root.config(bg=C["bg"])
    # 鎖定視窗大小(讓玩家放大縮小視窗，避免UI版面亂掉)
    root.resizable(False, False)

    # 設定下方的分頁選項
    NAV_H = 56
    # 建立一個名叫 content_area的Frame，用來裝所有核心功能畫面。
    content_area = tk.Frame(root, bg=C["bg"])
    content_area.place(x=0, y=0, width=W, height=H - NAV_H)

    # 設定遊戲需要的頁面
    timer_frame      = tk.Frame(content_area, bg=C["bg"])
    story_frame      = tk.Frame(content_area, bg=C["bg"])
    input_frame      = tk.Frame(content_area, bg=C["bg"])
    stat_frame       = tk.Frame(content_area, bg=C["bg"])
    story_list_frame = tk.Frame(content_area, bg=C["bg"])
    chat_frame       = tk.Frame(content_area, bg=C["bg"])
    route_frame      = tk.Frame(content_area, bg=C["bg"])

    ALL_FRAMES = [timer_frame, story_frame, input_frame,
                  stat_frame, story_list_frame, chat_frame, route_frame]

    # 顯示頁面的函式
    def show_frame(target):
        # 把所有視窗都關掉，再把目標視窗打開
        for f in ALL_FRAMES:
            f.pack_forget()
        target.pack(fill="both", expand=True)

    # 底部導覽列
    nav = tk.Frame(root, bg=C["surface"],
                   highlightbackground=C["border"],
                   highlightthickness=1)
    nav.place(x=0, y=H - NAV_H, width=W, height=NAV_H)
    nav_btns = {}

    def make_nav_btn(parent, label, icon, col, cmd):
        f = tk.Frame(parent, bg=C["surface"], cursor="hand2")
        f.grid(row=0, column=col, sticky="nsew")
        parent.columnconfigure(col, weight=1)
        il = tk.Label(f, text=icon, font=("Helvetica Neue", 16),
                      bg=C["surface"], fg=C["text3"])
        il.pack(pady=(8, 0))
        tl = tk.Label(f, text=label, font=("Helvetica Neue", 9),
                      bg=C["surface"], fg=C["text3"])
        tl.pack()
        for w in [f, il, tl]:
            w.bind("<Button-1>", lambda e, c=cmd: c())
        nav_btns[label] = (il, tl)

    def set_active_nav(label):
        for lbl, (il, tl) in nav_btns.items():
            color = C["pink"] if lbl == label else C["text3"]
            il.config(fg=color);  tl.config(fg=color)

    # ════════════════════════════════════════════
    #  頁面 1：計時器主頁
    # ════════════════════════════════════════════

    def build_timer_page():
        clear_frame(timer_frame)

        header = tk.Frame(timer_frame, bg=C["bg"])
        header.pack(fill="x", padx=20, pady=(16, 0))
        tk.Label(header, text="讀書計時器",
                 font=("Georgia", 20, "bold"),
                 bg=C["bg"], fg=C["text"]).pack(side="left")

        # 路線標籤
        if state.story_route:
            rc = ROUTE_CONFIG[state.story_route]
            tk.Label(header, text=rc["label"].split("  ")[1],
                     font=("Helvetica Neue", 10),
                     bg=rc["btn_bg"], fg=rc["btn_fg"],
                     padx=8, pady=3).pack(side="right")

        # EXP + 好感度卡
        status_bar = tk.Frame(timer_frame, bg=C["bg"])
        status_bar.pack(fill="x", padx=20, pady=(8, 0))

        exp_card = tk.Frame(status_bar, bg=C["gold_pale"],
                            highlightbackground="#E8C97A", highlightthickness=1)
        exp_card.pack(side="left", expand=True, fill="x", padx=(0, 6))
        exp_top = tk.Frame(exp_card, bg=C["gold_pale"])
        exp_top.pack(fill="x", padx=8, pady=(6, 2))
        tk.Label(exp_top, text="EXP", font=("Helvetica Neue", 9),
                 bg=C["gold_pale"], fg=C["gold"]).pack(side="left")
        exp_val_lbl = tk.Label(exp_top, text=f"{state.total_exp:.0f}",
                               font=("Helvetica Neue", 11, "bold"),
                               bg=C["gold_pale"], fg=C["gold"])
        exp_val_lbl.pack(side="right")
        exp_bar_canvas = tk.Canvas(exp_card, height=4, bg=C["border"], highlightthickness=0)
        exp_bar_canvas.pack(fill="x", padx=8, pady=(0, 2))
        exp_next_lbl = tk.Label(exp_card, text="",
                                font=("Helvetica Neue", 8),
                                bg=C["gold_pale"], fg=C["gold"])
        exp_next_lbl.pack(pady=(0, 4))

        aff_card = tk.Frame(status_bar, bg=C["pink_pale"],
                            highlightbackground="#F4C0D1", highlightthickness=1)
        aff_card.pack(side="left", expand=True, fill="x", padx=(6, 0))
        aff_top = tk.Frame(aff_card, bg=C["pink_pale"])
        aff_top.pack(fill="x", padx=8, pady=(6, 2))
        tk.Label(aff_top, text="好感度", font=("Helvetica Neue", 9),
                 bg=C["pink_pale"], fg=C["pink_text"]).pack(side="left")
        aff_val_lbl = tk.Label(aff_top, text=f"{state.affection:.0f}",
                               font=("Helvetica Neue", 11, "bold"),
                               bg=C["pink_pale"], fg=C["pink_text"])
        aff_val_lbl.pack(side="right")
        aff_bar_canvas = tk.Canvas(aff_card, height=4, bg=C["border"], highlightthickness=0)
        aff_bar_canvas.pack(fill="x", padx=8, pady=(0, 2))
        aff_level_lbl = tk.Label(aff_card,
                                 text=get_affection_level(state.affection)["name"],
                                 font=("Helvetica Neue", 8),
                                 bg=C["pink_pale"], fg=C["pink_text"])
        aff_level_lbl.pack(pady=(0, 4))

        def _draw_bar(canvas, pct, color):
            canvas.update_idletasks()
            w = canvas.winfo_width()
            if w <= 1: w = 160
            canvas.delete("bar")
            fill_w = max(0, int(w * pct))
            if fill_w > 0:
                canvas.create_rectangle(0, 0, fill_w, 4, fill=color, outline="", tags="bar")

        def refresh_status_bar():
            exp_val_lbl.config(text=f"{state.total_exp:.0f}")
            aff_val_lbl.config(text=f"{state.affection:.0f}")
            aff_level_lbl.config(text=get_affection_level(state.affection)["name"])
            next_story = next(
                (s for s in get_active_stories(state)
                 if not state.story_flags.get(s["id"]) and s["exp_threshold"] > 0),
                None)
            if next_story:
                pct    = min(state.total_exp / next_story["exp_threshold"], 1.0)
                remain = max(0, next_story["exp_threshold"] - state.total_exp)
                exp_next_lbl.config(text=f"下一章還差 {remain:.0f} EXP")
            else:
                pct = 1.0
                exp_next_lbl.config(text="所有章節已解鎖！")
            _draw_bar(exp_bar_canvas, pct, C["gold_bar"])
            _draw_bar(aff_bar_canvas, state.affection / AFF_MAX, C["pink_light"])

        timer_frame.after(100, refresh_status_bar)
        refresh_status_bar()

        # 進度環
        ring_size = 200
        ring_canvas = tk.Canvas(timer_frame, width=ring_size, height=ring_size,
                                bg=C["bg"], highlightthickness=0)
        ring_canvas.pack(pady=(6, 0))
        elapsed_lbl = tk.Label(ring_canvas, text="00:00:00",
                               font=("Georgia", 24, "bold"), bg=C["bg"], fg=C["text"])
        elapsed_lbl.place(relx=0.5, rely=0.40, anchor="center")
        tk.Label(ring_canvas, text="目前連續時間",
                 font=("Helvetica Neue", 10), bg=C["bg"], fg=C["text3"]
                 ).place(relx=0.5, rely=0.56, anchor="center")
        total_ring_lbl = tk.Label(ring_canvas, text=f"總計 {fmt(state.total_time)}",
                                  font=("Helvetica Neue", 10), bg=C["bg"], fg=C["text2"])
        total_ring_lbl.place(relx=0.5, rely=0.68, anchor="center")
        cx, cy, r = ring_size // 2, ring_size // 2, 80

        def redraw_ring():
            ring_canvas.delete("ring")
            nxt = next((s["exp_threshold"] for s in get_active_stories(state)
                        if s["exp_threshold"] > 0 and not state.story_flags.get(s["id"])),
                       None)
            pct = min(state.total_exp / nxt, 1.0) if nxt else 1.0
            ring_canvas.create_arc(cx-r, cy-r, cx+r, cy+r, start=0, extent=359.99,
                                   outline=C["border"], width=12, style="arc", tags="ring")
            if pct > 0:
                ring_canvas.create_arc(cx-r, cy-r, cx+r, cy+r, start=90,
                                       extent=-(pct*359.99), outline=C["gold_bar"],
                                       width=12, style="arc", tags="ring")
            if pct > 0.01:
                angle = math.radians(90 - pct * 360)
                tx = cx + r * math.cos(angle);  ty = cy - r * math.sin(angle)
                ring_canvas.create_oval(tx-6, ty-6, tx+6, ty+6,
                                        fill=C["gold"], outline="", tags="ring")

        # 統計小卡列
        stat_row = tk.Frame(timer_frame, bg=C["bg"])
        stat_row.pack(fill="x", padx=20, pady=(4, 0))

        def make_mini_card(parent, label_text, value_text):
            card = tk.Frame(parent, bg=C["surface"],
                            highlightbackground=C["border"], highlightthickness=1)
            card.pack(side="left", expand=True, fill="x", padx=3)
            tk.Label(card, text=label_text, font=("Helvetica Neue", 9),
                     bg=C["surface"], fg=C["text3"]).pack(pady=(6, 0))
            lbl = tk.Label(card, text=value_text,
                           font=("Helvetica Neue", 12, "bold"),
                           bg=C["surface"], fg=C["text"])
            lbl.pack(pady=(1, 6))
            return lbl

        count_lbl = make_mini_card(stat_row, "計時次數", f"{state.session_count} 次")
        avg_lbl   = make_mini_card(stat_row, "平均時長", fmt_min(state.average_time))
        total_lbl = make_mini_card(stat_row, "累積時間", fmt(state.total_time))

        # 按鈕列
        btn_row = tk.Frame(timer_frame, bg=C["bg"])
        btn_row.pack(fill="x", padx=20, pady=(10, 0))
        tk.Button(btn_row, text="▶  開始",
                  font=("Helvetica Neue", 13, "bold"),
                  bg=C["pink"], fg="white", activebackground=C["pink_text"],
                  relief="flat", bd=0, cursor="hand2", padx=0, pady=9,
                  command=lambda: start_timer(state)
                  ).pack(side="left", expand=True, fill="x", padx=(0, 5))
        tk.Button(btn_row, text="⏸  暫停",
                  font=("Helvetica Neue", 13, "bold"),
                  bg=C["surface"], fg=C["text"], activebackground=C["surface2"],
                  relief="flat", bd=0, cursor="hand2",
                  highlightbackground=C["border"], highlightthickness=1,
                  padx=0, pady=9,
                  command=lambda: on_pause()
                  ).pack(side="left", expand=True, fill="x", padx=(5, 0))

        # 提醒
        notice = tk.Frame(timer_frame, bg=C["pink_pale"],
                          highlightbackground="#F4C0D1", highlightthickness=1)
        notice.pack(fill="x", padx=20, pady=(8, 0))
        tk.Label(notice, text="暫停後才會儲存紀錄，關閉前請記得暫停喔！",
                 font=("Helvetica Neue", 10),
                 bg=C["pink_pale"], fg=C["pink_text"],
                 wraplength=340, pady=6).pack()

        ## 劇情 mini strip（已移除）
        def refresh_story_strip():
            pass

        def update_time():
            if state.running:
                state.elapsed_time = time.time() - state.start_time
                elapsed_lbl.config(text=fmt(state.elapsed_time))
                total_ring_lbl.config(text=f"總計 {fmt(state.total_time + state.elapsed_time)}")
            redraw_ring()
            root.after(1000, update_time)

        update_time()

        def on_pause():
            do_pause(state, {"on_pause": _after_pause, "on_story": trigger_story},
                     planned_secs=0, distraction_count=0)

        def _after_pause(exp_gained, aff_gained):
            elapsed_lbl.config(text="00:00:00")
            total_ring_lbl.config(text=f"總計 {fmt(state.total_time)}")
            count_lbl.config(text=f"{state.session_count} 次")
            avg_lbl.config(text=fmt_min(state.average_time))
            total_lbl.config(text=fmt(state.total_time))
            refresh_status_bar()
            refresh_story_strip()
            toast = tk.Label(timer_frame,
                             text=f"+{exp_gained:.1f} EXP　+{aff_gained:.1f} 好感度",
                             font=("Helvetica Neue", 11, "bold"),
                             bg=C["teal_pale"], fg=C["teal_text"],
                             padx=14, pady=6,
                             highlightbackground=C["teal"], highlightthickness=1)
            toast.place(relx=0.5, rely=0.97, anchor="s")
            timer_frame.after(2500, toast.destroy)

        story_queue = []

        def play_next_story():
            if not story_queue:
                refresh_story_strip()
                refresh_status_bar()
                show_frame(timer_frame)
                return

            story_def   = story_queue.pop(0)
            parent_id   = story_def.get("parent_id", story_def["id"])
            parent_cell = story_def["excel_cell"]
            branch_id   = story_def.get("branch_id")
            branch_cell = story_def.get("branch_cell")

            def on_finished():
                state.story_flags[parent_id] = True
                save_story_flag(state.game_data, parent_cell, True)
                if branch_id and branch_cell:
                    state.story_flags[branch_id] = True
                    save_story_flag(state.game_data, branch_cell, True)
                play_next_story()

            show_story(story_def, state.user_name,
                       story_frame, timer_frame,
                       route=state.story_route,
                       on_finished=on_finished)

        def trigger_story(story_def):
            story_queue.append(story_def)
            if len(story_queue) == 1:
                play_next_story()

        def trigger_story0():
            """觸發共用序章。"""
            s0 = STORY_PROLOGUE
            def on_finished():
                state.story_flags[s0["id"]] = True
                save_story_flag(state.game_data, s0["excel_cell"], True)
                _after_prologue()
            def on_no_finish():
                save_story_flag(state.game_data, s0["excel_cell"], False)
            show_story(s0, state.user_name, story_frame, timer_frame,
                       route="romance",   # 序章用乙女向風格（共用）
                       on_no_finish=on_no_finish, on_finished=on_finished)

        def _after_prologue():
            """序章結束後：輸入名字 → 選路線 → 計時器。"""
            if not state.user_name:
                build_input_page(back_fn=lambda: _after_name())
                show_frame(input_frame)
            else:
                _after_name()

        def _after_name():
            if not state.story_route:
                build_route_select_page(back_fn=lambda: show_frame(timer_frame))
                show_frame(route_frame)
            else:
                show_frame(timer_frame)

        return trigger_story, trigger_story0, _after_prologue, _after_name

    # ════════════════════════════════════════════
    #  頁面 2：輸入名字
    # ════════════════════════════════════════════

    def build_input_page(back_fn=None):
        clear_frame(input_frame)
        tk.Label(input_frame, text="請輸入你的名字",
                 font=("Georgia", 20, "bold"),
                 bg=C["bg"], fg=C["text"]).pack(pady=(80, 6))
        tk.Label(input_frame, text="名字會出現在劇情對話中",
                 font=("Helvetica Neue", 12), bg=C["bg"], fg=C["text3"]).pack(pady=(0, 30))
        name_var = tk.StringVar()
        entry = tk.Entry(input_frame, textvariable=name_var,
                         font=("Helvetica Neue", 16), relief="flat", bd=0,
                         highlightbackground=C["border2"], highlightthickness=1,
                         bg=C["surface"], fg=C["text"], insertbackground=C["text"],
                         justify="center")
        entry.pack(padx=50, ipady=10, fill="x")
        entry.focus()
        def confirm():
            name = name_var.get().strip()
            if not name: return
            state.user_name = name
            save_user_name(state.game_data, name)
            if back_fn: back_fn()
        tk.Button(input_frame, text="確認",
                  font=("Helvetica Neue", 14, "bold"),
                  bg=C["pink"], fg="white", activebackground=C["pink_text"],
                  relief="flat", bd=0, cursor="hand2", padx=0, pady=10,
                  command=confirm).pack(padx=50, pady=(16, 0), fill="x")
        entry.bind("<Return>", lambda e: confirm())

    # ════════════════════════════════════════════
    #  頁面 3：故事線選擇
    # ════════════════════════════════════════════

    def build_route_select_page(back_fn=None):
        clear_frame(route_frame)

        tk.Label(route_frame, text="選擇你的故事",
                 font=("Georgia", 24, "bold"),
                 bg=C["bg"], fg=C["text"]).pack(pady=(60, 6))
        tk.Label(route_frame, text="選擇後將開始解鎖對應劇情",
                 font=("Helvetica Neue", 12),
                 bg=C["bg"], fg=C["text3"]).pack(pady=(0, 36))

        for route_key, rc in ROUTE_CONFIG.items():
            card = tk.Frame(route_frame, bg=rc["btn_bg"],
                            highlightbackground=C["border2"],
                            highlightthickness=1)
            card.pack(fill="x", padx=40, pady=(0, 16))

            tk.Label(card, text=rc["label"],
                     font=("Helvetica Neue", 16, "bold"),
                     bg=rc["btn_bg"], fg=rc["btn_fg"]).pack(pady=(18, 4))
            tk.Label(card, text=rc["description"],
                     font=("Helvetica Neue", 11),
                     bg=rc["btn_bg"], fg=rc["btn_fg"],
                     justify="center", wraplength=280).pack(pady=(0, 6))

            btn_text = "選擇此故事線 →"
            btn = tk.Button(card, text=btn_text,
                            font=("Helvetica Neue", 12),
                            bg=C["surface"], fg=rc["btn_bg"],
                            activebackground=C["surface2"],
                            relief="flat", bd=0, cursor="hand2",
                            padx=14, pady=6)
            btn.pack(pady=(4, 18))

            def make_choose(rk=route_key):
                def choose():
                    state.story_route = rk
                    save_story_route(state.game_data, rk)
                    if back_fn: back_fn()
                return choose

            btn.config(command=make_choose())

    # ════════════════════════════════════════════
    #  頁面 4：統計頁
    # ════════════════════════════════════════════

    def build_stat_page():
        clear_frame(stat_frame)
        tk.Label(stat_frame, text="讀書統計",
                 font=("Georgia", 18, "bold"),
                 bg=C["bg"], fg=C["text"]).pack(anchor="w", padx=20, pady=(16, 0))
        make_sep(stat_frame).pack(fill="x", padx=20, pady=(10, 0))

        grid = tk.Frame(stat_frame, bg=C["bg"])
        grid.pack(fill="x", padx=20, pady=(12, 0))

        def big_card(parent, row, col, label, value, unit=""):
            card = tk.Frame(parent, bg=C["surface"],
                            highlightbackground=C["border"], highlightthickness=1)
            card.grid(row=row, column=col, padx=5, pady=5, sticky="nsew")
            tk.Label(card, text=label, font=("Helvetica Neue", 10),
                     bg=C["surface"], fg=C["text3"]).pack(pady=(10, 0))
            vf = tk.Frame(card, bg=C["surface"])
            vf.pack(pady=(2, 10))
            tk.Label(vf, text=value, font=("Georgia", 17, "bold"),
                     bg=C["surface"], fg=C["text"]).pack(side="left")
            if unit:
                tk.Label(vf, text=f" {unit}", font=("Helvetica Neue", 10),
                         bg=C["surface"], fg=C["text3"]).pack(side="left", anchor="s", pady=(0, 3))

        grid.columnconfigure(0, weight=1);  grid.columnconfigure(1, weight=1)
        h_total = int(state.total_time // 3600)
        m_total = int((state.total_time % 3600) // 60)
        total_str = f"{h_total}h {m_total:02d}m" if h_total else f"{m_total} 分"
        avg_m = int(state.average_time // 60);  avg_s = int(state.average_time % 60)
        sessions = load_session_history(state.game_data)
        max_m = max((time_str_to_sec(t) for t in sessions), default=0) // 60
        big_card(grid, 0, 0, "累積讀書時間", total_str)
        big_card(grid, 0, 1, "平均每次時長", f"{avg_m}:{avg_s:02d}", "分鐘")
        big_card(grid, 1, 0, "計時次數", str(state.session_count), "次")
        big_card(grid, 1, 1, "最長單次", str(max_m), "分鐘")

        num_row = tk.Frame(stat_frame, bg=C["bg"])
        num_row.pack(fill="x", padx=20)
        for label, value, unit, bg, fg in [
            ("累積 EXP", f"{state.total_exp:.0f}", "pts", C["gold_pale"], C["gold"]),
            (f"好感度（{get_affection_level(state.affection)['name']}）",
             f"{state.affection:.0f}", "/ 100", C["pink_pale"], C["pink_text"]),
        ]:
            card = tk.Frame(num_row, bg=bg, highlightbackground=fg, highlightthickness=1)
            card.pack(side="left", expand=True, fill="x", padx=5, pady=5)
            tk.Label(card, text=label, font=("Helvetica Neue", 10), bg=bg, fg=fg).pack(pady=(8, 0))
            vf = tk.Frame(card, bg=bg); vf.pack(pady=(2, 8))
            tk.Label(vf, text=value, font=("Georgia", 17, "bold"), bg=bg, fg=fg).pack(side="left")
            tk.Label(vf, text=f" {unit}", font=("Helvetica Neue", 10), bg=bg, fg=fg).pack(
                side="left", anchor="s", pady=(0, 2))

        tk.Label(stat_frame, text="最近 5 次讀書時長",
                 font=("Helvetica Neue", 11), bg=C["bg"], fg=C["text3"]
                 ).pack(anchor="w", padx=24, pady=(10, 4))
        chart_card = tk.Frame(stat_frame, bg=C["surface"],
                              highlightbackground=C["border"], highlightthickness=1)
        chart_card.pack(fill="x", padx=20)
        fig, ax = plt.subplots(figsize=(4.5, 1.8), dpi=90)
        fig.patch.set_facecolor(C["surface"]); ax.set_facecolor(C["surface"])
        last5 = sessions[-5:] if sessions else []
        if last5:
            mins = [time_str_to_min(t) for t in last5]
            x = range(1, len(mins) + 1)
            ax.bar(x, mins,
                   color=[C["pink"] if i == len(mins)-1 else C["pink_light"] for i in range(len(mins))],
                   width=0.5, zorder=3)
            ax.set_ylim(0, max(mins) * 1.3)
            ax.set_xticks(list(x)); ax.set_xticklabels([str(i) for i in x], fontsize=9, color=C["text3"])
            ax.tick_params(axis="y", labelsize=9, labelcolor=C["text3"])
            for sp in ax.spines.values(): sp.set_visible(False)
            ax.yaxis.set_tick_params(length=0); ax.xaxis.set_tick_params(length=0)
            ax.grid(axis="y", color=C["border"], linewidth=0.7, zorder=0)
        else:
            ax.text(0.5, 0.5, "還沒有紀錄", ha="center", va="center",
                    transform=ax.transAxes, color=C["text3"], fontsize=11)
            for sp in ax.spines.values(): sp.set_visible(False)
            ax.set_xticks([]); ax.set_yticks([])
        fig.tight_layout(pad=0.8)
        canvas = FigureCanvasTkAgg(fig, master=chart_card)
        canvas.draw(); canvas.get_tk_widget().pack(fill="x")
        plt.close(fig)

    # ════════════════════════════════════════════
    #  頁面 5：劇情進度頁
    # ════════════════════════════════════════════

    def build_story_list_page():
        clear_frame(story_list_frame)
        tk.Label(story_list_frame, text="劇情進度",
                 font=("Georgia", 18, "bold"),
                 bg=C["bg"], fg=C["text"]).pack(anchor="w", padx=20, pady=(16, 0))
        make_sep(story_list_frame).pack(fill="x", padx=20, pady=(10, 0))

        stories = get_active_stories(state)
        if not stories:
            tk.Label(story_list_frame, text="尚未選擇故事線",
                     font=("Helvetica Neue", 13),
                     bg=C["bg"], fg=C["text3"]).pack(pady=60)
            return

        done_count  = sum(1 for s in stories if state.story_flags.get(s["id"]))
        total_count = len(stories)
        route_name  = ROUTE_CONFIG.get(state.story_route, {}).get("label", "")

        tk.Label(story_list_frame,
                 text=f"{route_name}  ·  已解鎖 {done_count} / {total_count} 章",
                 font=("Helvetica Neue", 12),
                 bg=C["bg"], fg=C["text2"]).pack(anchor="w", padx=24, pady=(8, 2))
        prog_bg = tk.Frame(story_list_frame, bg=C["border"], height=4)
        prog_bg.pack(fill="x", padx=20, pady=(0, 12))
        tk.Frame(prog_bg, bg=C["pink"], height=4).place(
            relwidth=done_count / total_count if total_count else 0, relheight=1)

        outer = tk.Frame(story_list_frame, bg=C["bg"])
        outer.pack(fill="both", expand=True, padx=20)
        cs = tk.Canvas(outer, bg=C["bg"], highlightthickness=0)
        sb = tk.Scrollbar(outer, orient="vertical", command=cs.yview)
        cs.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y"); cs.pack(side="left", fill="both", expand=True)
        inner = tk.Frame(cs, bg=C["bg"])
        cs.create_window((0, 0), window=inner, anchor="nw")
        inner.bind("<Configure>", lambda e: cs.configure(scrollregion=cs.bbox("all")))

        for s in stories:
            sid  = s["id"]
            done = state.story_flags.get(sid, False)
            card = tk.Frame(inner, bg=C["surface"],
                            highlightbackground=C["border"], highlightthickness=1)
            card.pack(fill="x", pady=(0, 10))
            tk.Frame(card, bg=C["teal"] if done else C["border2"], width=4).pack(side="left", fill="y")
            body = tk.Frame(card, bg=C["surface"])
            body.pack(side="left", fill="both", expand=True, padx=14, pady=12)

            title_row = tk.Frame(body, bg=C["surface"])
            title_row.pack(fill="x")
            tk.Label(title_row, text=s["title"], font=("Helvetica Neue", 10),
                     bg=C["surface"], fg=C["teal_text"] if done else C["text3"]).pack(side="left")
            status_text = ("✓ 已解鎖" if done else
                           ("🔒 未解鎖" if s["exp_threshold"] > 0 else "🔓 可解鎖"))
            tk.Label(title_row, text=status_text, font=("Helvetica Neue", 10),
                     bg=C["surface"], fg=C["teal"] if done else C["gray"]).pack(side="right")

            # 副標題（有分歧時顯示父章節 subtitle，避免暴雷）
            display_subtitle = s["subtitle"]
            tk.Label(body, text=display_subtitle, font=("Georgia", 14, "bold"),
                     bg=C["surface"], fg=C["text"] if done else C["text3"]
                     ).pack(anchor="w", pady=(2, 0))

            if not done and s["exp_threshold"] > 0:
                need   = s["exp_threshold"]
                cur    = min(state.total_exp, need)
                pct    = cur / need
                remain = max(0, need - state.total_exp)
                tk.Label(body, text=f"還差 {remain:.0f} EXP 解鎖（累積 {cur:.0f} / {need:.0f}）",
                         font=("Helvetica Neue", 10),
                         bg=C["surface"], fg=C["gold"]).pack(anchor="w", pady=(4, 2))
                bar_bg2 = tk.Frame(body, bg=C["gold_pale"], height=4)
                bar_bg2.pack(fill="x", pady=(0, 2))
                tk.Frame(bar_bg2, bg=C["gold_bar"], height=4).place(relwidth=pct, relheight=1)

            req = s.get("requires")
            if not done and req and not state.story_flags.get(req):
                req_s = next((x for x in stories if x["id"] == req), None)
                if req_s:
                    tk.Label(body, text=f"需先解鎖「{req_s['title']}」",
                             font=("Helvetica Neue", 9),
                             bg=C["surface"], fg=C["text3"]).pack(anchor="w")

            if done:
                def make_replay(story_def=s):
                    def replay():
                        if "branches" in story_def:
                        # 找玩家實際看過的支線（story_flags 有記錄的那條）
                            seen_branch = next(
                            (b for b in story_def["branches"]
                            if state.story_flags.get(b["id"])),
                            None
                            )
                        # 找不到（沒有 branch_cell 的舊資料）才 fallback 到條件判斷
                            if seen_branch:
                                target = {**story_def, **seen_branch}
                            else:
                                chosen = pick_branch(story_def["branches"], state)
                                target = {**story_def, **chosen}
                        else:
                            target = story_def

                        show_story(target, state.user_name,
                        story_frame, story_list_frame,
                        route=state.story_route,
                        on_finished=lambda: show_frame(story_list_frame))
                        show_frame(story_frame)
                    return replay
            
                tk.Button(body, text="複習劇情",
                          font=("Helvetica Neue", 11),
                          bg=C["pink_pale"], fg=C["pink_text"],
                          activebackground="#F4C0D1", relief="flat", bd=0, cursor="hand2",
                          highlightbackground="#F4C0D1", highlightthickness=1,
                          padx=12, pady=5,
                          command=make_replay()).pack(anchor="w", pady=(8, 0))

        cs.bind_all("<MouseWheel>", lambda e: cs.yview_scroll(int(-1*(e.delta/120)), "units"))

    # ════════════════════════════════════════════
    #  頁面 6：日常劇情聊天頁
    # ════════════════════════════════════════════

    def build_chat_page():
        clear_frame(chat_frame)
        header = tk.Frame(chat_frame, bg=C["surface"],
                          highlightbackground=C["border"], highlightthickness=1)
        header.pack(fill="x")
        tk.Label(header, text="林霽安", font=("Georgia", 15, "bold"),
                 bg=C["surface"], fg=C["text"]).pack(side="left", padx=16, pady=12)
        lvl = get_affection_level(state.affection)
        tk.Label(header, text=f"♥ {lvl['name']}　好感度 {state.affection:.0f}",
                 font=("Helvetica Neue", 10),
                 bg=C["surface"], fg=C["pink_text"]).pack(side="right", padx=16)

        chat_outer = tk.Frame(chat_frame, bg=C["surface2"])
        chat_outer.pack(fill="both", expand=True)
        chat_canvas = tk.Canvas(chat_outer, bg=C["surface2"], highlightthickness=0)
        chat_sb = tk.Scrollbar(chat_outer, orient="vertical", command=chat_canvas.yview)
        chat_canvas.configure(yscrollcommand=chat_sb.set)
        chat_sb.pack(side="right", fill="y"); chat_canvas.pack(side="left", fill="both", expand=True)

        msg_area  = tk.Frame(chat_canvas, bg=C["surface2"])
        msg_win   = chat_canvas.create_window((0, 0), window=msg_area, anchor="nw")
        msg_area.bind("<Configure>", lambda e: (
            chat_canvas.configure(scrollregion=chat_canvas.bbox("all")),
            chat_canvas.itemconfig(msg_win, width=chat_canvas.winfo_width())))
        chat_canvas.bind("<Configure>",
                         lambda e: chat_canvas.itemconfig(msg_win, width=e.width))

        def add_char_bubble(text):
            row = tk.Frame(msg_area, bg=C["surface2"])
            row.pack(fill="x", padx=12, pady=(6, 0), anchor="w")
            tk.Label(row, text="霽", font=("Georgia", 11, "bold"),
                     bg=C["pink_light"], fg="white",
                     width=2, height=1, padx=4, pady=4).pack(side="left", anchor="n")
            tk.Label(row, text=text, font=("Helvetica Neue", 12),
                     bg=C["bubble_char"], fg=C["text"],
                     wraplength=240, justify="left", padx=12, pady=8,
                     highlightbackground=C["border"], highlightthickness=1
                     ).pack(side="left", padx=(6, 0))

        def add_date_divider(text):
            row = tk.Frame(msg_area, bg=C["surface2"])
            row.pack(fill="x", padx=20, pady=(12, 4))
            tk.Frame(row, bg=C["border"], height=1).pack(side="left", expand=True, fill="x")
            tk.Label(row, text=f"  {text}  ", font=("Helvetica Neue", 9),
                     bg=C["surface2"], fg=C["text3"]).pack(side="left")
            tk.Frame(row, bg=C["border"], height=1).pack(side="left", expand=True, fill="x")

        def scroll_to_bottom():
            chat_canvas.update_idletasks(); chat_canvas.yview_moveto(1.0)

        add_date_divider("今天")
        lvl_name = get_affection_level(state.affection)["name"]
        pool = DAILY_CHATS.get(lvl_name, DAILY_CHATS["陌生"])
        random.seed(int(today_str().replace("-", "")))
        chosen = random.choice(pool)

        def show_messages(messages, idx=0):
            if idx >= len(messages):
                if not state.today_chat_shown:
                    state.today_chat_shown = True
                    save_today_chat_shown(state.game_data)
                scroll_to_bottom(); return
            add_char_bubble(messages[idx]); scroll_to_bottom()
            chat_frame.after(800, lambda: show_messages(messages, idx + 1))

        show_messages(chosen)

        input_bar = tk.Frame(chat_frame, bg=C["surface"],
                             highlightbackground=C["border"], highlightthickness=1)
        input_bar.pack(fill="x", side="bottom")
        tk.Label(input_bar, text="選項回覆功能開發中…",
                 font=("Helvetica Neue", 11), bg=C["surface"], fg=C["text3"], pady=12).pack()

    # ════════════════════════════════════════════
    #  導覽列 & 啟動
    # ════════════════════════════════════════════

    def open_timer():
        set_active_nav("計時"); show_frame(timer_frame)

    def open_chat():
        set_active_nav("日常"); build_chat_page(); show_frame(chat_frame)

    def open_stat():
        if state.running: return
        set_active_nav("統計"); build_stat_page(); show_frame(stat_frame)

    def open_story_list():
        set_active_nav("劇情"); build_story_list_page(); show_frame(story_list_frame)

    make_nav_btn(nav, "計時", "⏱", 0, open_timer)
    make_nav_btn(nav, "日常", "💬", 1, open_chat)
    make_nav_btn(nav, "統計", "📊", 2, open_stat)
    make_nav_btn(nav, "劇情", "📖", 3, open_story_list)

    trigger_story, trigger_story0, _after_prologue, _after_name = build_timer_page()
    set_active_nav("計時")

    # 好感度衰減提示
    if decay_amount > 0:
        def show_decay_notice():
            toast = tk.Label(content_area,
                             text=f"好感度因為你的缺席下降了 {decay_amount:.0f} 點…",
                             font=("Helvetica Neue", 11),
                             bg="#FFF0F5", fg=C["pink_text"],
                             padx=14, pady=8,
                             highlightbackground=C["pink_light"], highlightthickness=1,
                             wraplength=300)
            toast.place(relx=0.5, rely=0.1, anchor="n")
            content_area.after(3000, toast.destroy)
        root.after(500, show_decay_notice)

    # 啟動流程
    prologue_done = state.story_flags.get(STORY_PROLOGUE["id"], False)

    if not prologue_done:
        # 全新玩家：跑序章 → 輸入名字 → 選路線
        trigger_story0()
    elif not state.user_name:
        # 序章看了但沒輸入名字（中途關閉）
        build_input_page(back_fn=lambda: _after_name())
        show_frame(input_frame)
    elif not state.story_route:
        # 名字輸入了但沒選路線（中途關閉）
        build_route_select_page(back_fn=lambda: show_frame(timer_frame))
        show_frame(route_frame)
    else:
        # 正常重開：補觸發未完成的章節
        check_story_triggers(state, on_story=trigger_story)
        show_frame(timer_frame)

    root.mainloop()


if __name__ == "__main__":
    main()
