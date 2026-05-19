# 讀書計時器 — 數值系統版
# 新增：EXP 計算、好感度計算與衰減、日常劇情聊天頁面
# 主線劇情改以 EXP 為解鎖門檻

import subprocess
import sys
from dataclasses import dataclass, field

required_packages = ["openpyxl", "matplotlib"]
for package in required_packages:
    try:
        __import__(package)
    except ImportError:
        print(f"安裝 {package} 中，請稍等...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", package])

import tkinter as tk
import time
import os
import math
import datetime
from openpyxl import Workbook, load_workbook
import matplotlib
matplotlib.use("TkAgg")
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg


# ════════════════════════════════════════════════
#  設計系統 (Design Tokens)
# ════════════════════════════════════════════════

C = {
    "bg":         "#FAFAF8",
    "surface":    "#FFFFFF",
    "surface2":   "#F5F3EE",
    "border":     "#E8E4DE",
    "border2":    "#D5CFC6",
    "pink":       "#D4537E",
    "pink_light": "#ED93B1",
    "pink_pale":  "#FBEAF0",
    "pink_text":  "#993556",
    "teal":       "#1D9E75",
    "teal_pale":  "#E1F5EE",
    "teal_text":  "#0F6E56",
    "gold":       "#C8922A",
    "gold_pale":  "#FDF3E0",
    "gold_bar":   "#E8B84B",
    "gray":       "#888780",
    "gray_pale":  "#F1EFE8",
    "gray_text":  "#5F5E5A",
    "text":       "#2C2C2A",
    "text2":      "#5F5E5A",
    "text3":      "#888780",
    "bubble_char":"#FFF0F5",
    "bubble_user":"#D4537E",
}

W, H = 400, 680


# ════════════════════════════════════════════════
#  好感度階層定義
# ════════════════════════════════════════════════

AFFECTION_LEVELS = [
    {"threshold": 0,  "name": "陌生",   "desc": "對方對你還不太熟悉……"},
    {"threshold": 20, "name": "禮貌",   "desc": "對方以禮相待，但仍保持距離。"},
    {"threshold": 40, "name": "親近",   "desc": "對方開始對你展露笑容。"},
    {"threshold": 60, "name": "信任",   "desc": "對方願意和你分享心事了。"},
    {"threshold": 80, "name": "親密",   "desc": "對方在你身邊感到很自在。"},
]

def get_affection_level(aff: float) -> dict:
    current = AFFECTION_LEVELS[0]
    for lvl in AFFECTION_LEVELS:
        if aff >= lvl["threshold"]:
            current = lvl
    return current


# ════════════════════════════════════════════════
#  日常劇情資料
#  新增日常劇情：在對應階層的 list 中加一個 list 即可
# ════════════════════════════════════════════════

DAILY_CHATS = {
    "陌生": [
        ["今天有讀書嗎？", "……保持習慣很重要。"],
        ["嗯。", "加油。"],
    ],
    "禮貌": [
        ["你今天讀了多久？", "不管多少，有讀就好。繼續努力。"],
        ["讀書辛苦了。", "記得休息。"],
        ["今天狀況還好嗎？", "學習要循序漸進，不用太急。"],
    ],
    "親近": [
        ["今天讀得怎麼樣？", "有沒有遇到什麼難題？"],
        ["你最近讀書的進度感覺不錯。", "繼續保持，我有在看的。"],
        ["……其實我也想多讀一點書。", "你能持續下去的話，讓我有點佩服你。"],
    ],
    "信任": [
        ["你知道嗎，我以前也有段時間完全讀不下去……", "後來是每天只逼自己讀十分鐘。", "你比我當時強多了。"],
        ["今天讀書了嗎。", "……我知道你會的。", "不知道為什麼，就是覺得你不會放棄。"],
        ["你有沒有想過，為什麼要這麼努力？", "我最近一直在想這個問題。", "……不用現在回答我，只是想說說。"],
    ],
    "親密": [
        ["今天也讀書了。", "……你真的很讓我安心。", "不知道為什麼這樣說，但就是這個感覺。"],
        ["嗯，我一直都有注意到你在努力。", "你知道嗎，每次看到你認真的樣子……", "算了，說出來會很奇怪。"],
        ["今天可以多讀一下嗎。", "我想多陪你一會兒。", "……這樣說很奇怪嗎？"],
    ],
}


# ════════════════════════════════════════════════
#  主線劇情資料
#  exp_threshold: 解鎖所需 EXP
# ════════════════════════════════════════════════

STORIES = [
    {
        "id": "story0",
        "title": "序章",
        "subtitle": "圖書館的午後",
        "exp_threshold": 0,
        "requires": None,
        "excel_cell": "H2",
        "text": [
            "點擊左鍵繼續劇情",
            "學期初的下午時分，我獨自一人來到圖書館的討論間中。",
            "輕輕地將門帶上後，我忍不住緊張地嘆了口氣。",
            "（早知道就不要一時衝動的參加這種活動了。。。）",
            "因為最近一直提不起勁讀書，我報名了學校舉辦的讀書會活動。",
            "據說，活動方會將參加者分組，並提供場地，讓大家舉行一場小型讀書會。",
            "成員們可以藉此互相認識、互相督促，如果彼此合得來，還能組成長期讀書小組，一起進步。",
            "（說是這麼說，但是果然還是有點尷尬阿。。。）",
            "（不過，都已經來了，現在後悔也來不及了。）",
            "我看著空蕩蕩的討論室，有些無奈地選了個最角落的位置坐下。",
            "瞥了眼手機上的時間，離讀書會開始還有四十分鐘左右。",
            "（看來是我來得太早了。）",
            "寧靜的討論室裡，午後溫暖的陽光透過大片窗戶，斜斜地灑落在光滑的木質桌面上，映出一片柔和的金色光暈。",
            "空氣中瀰漫著淡淡的木香與書頁的紙墨氣息。",
            "時鐘滴答作響，每一下聲音都清晰可聞，和窗外偶爾傳來的鳥鳴聲交織在一起，編織出一種溫柔又專注的靜謐氛圍。",
            "（這種氛圍真適合讀書呢。）",
            "「嘛，竟然還有這麼多時間，不如先讀一下書吧！」",
        ],
    },
    {
        "id": "story1",
        "title": "第一章",
        "subtitle": "讀書會成員",
        "exp_threshold": 100,
        "requires": None,
        "excel_cell": "H3",
        "text_fn": lambda name: [
            "「嘎吱──」",
            "伴隨著門軸緩慢轉動的聲響，討論室的門被輕輕推開。",
            "一位氣質內斂的青年踏了進來。他有著一頭微捲的黑色短髮，佩戴細框眼鏡，眼下浮著淡淡的黑眼圈。",
            "我們的視線在空中短暫交會，他像是被燙到似地立刻移開目光，略微側過頭，聲音很輕。",
            "？？？：「……你好。」",
            "「啊……你好。」",
            "林霽安：「我叫林霽安。」",
            f"「我叫{name}。」",
            "（……果然很尷尬！）",
            "正當我們對視又各自移開、不知該說些什麼時，討論室的門再次被打開。",
            "走進來的是一位穿著寬鬆運動T-shirt、身材高挑的男生，他的臉上掛著一抹耀眼的笑容。",
            "他身旁則是一位穿著深色高領毛衣的男子，步伐從容、氣質溫雅。",
            "？？？：「嗨！看起來人到齊了耶！」",
            "顧以恆：「我叫顧以恆，心理系二年級。能告訴我妳的名字嗎？」",
            "林霽安：「……宋曜陽，圖書館內禁止奔跑。」",
            "宋曜陽：「啊啊，抱歉啦～習慣性反應，太興奮了哈哈。」",
            "「那個……你們三個，難道早就彼此認識了嗎？」",
        ],
    },
    {
        "id": "story2",
        "title": "第二章",
        "subtitle": "不確定的相遇",
        "exp_threshold": 250,
        "requires": "story1",
        "excel_cell": "H4",
        "text_fn": lambda name: [
            "顧以恆依舊維持著那得體而溫和的笑容，將下巴輕輕倚在掌心，微微偏頭望向我。",
            "顧以恆：「嗯，我們是高中同學。」",
            "宋曜陽已笑嘻嘻地伸手摟住林霽安的肩，動作大方又毫無顧忌。",
            "宋曜陽：「嘿嘿～我們三個還約好一起報名這個活動的～」",
            "林霽安：「……不要碰我。」",
            "宋曜陽吐了吐舌頭，縮回手，語氣倒也沒有惱怒，反而笑得更加開懷。",
            "我看著眼前三人的互動，有些訝異又有些羨慕。三種截然不同的性格，卻意外地搭配得那麼自然。",
            f"顧以恆：「那，我們就一個個來正式自我介紹吧，也讓{name}不會覺得自己像個外人。」",
            "林霽安：「……林霽安，醫學系三年級。」",
            "宋曜陽：「我是宋曜陽，運動科學系大三！興趣是打籃球、吃宵夜、還有……帶讀書會氣氛！」",
            "顧以恆：「顧以恆，心理系二年級。興趣是閱讀跟觀察人。」",
            f"「我叫{name}，因為最近讀書動力有點低迷，就想說來參加看看……」",
            "宋曜陽：「我懂我懂，有時候一個人真的很難撐下去！」",
            "顧以恆：「讀書會的意義也正在於此，互相陪伴、互相督促。」",
            "林霽安：「……希望妳能專心。」",
            "我心頭一暖，忍不住勾起了嘴角。",
            "宋曜陽：「好了，那就開始今天的讀書會吧──GO！」",
            "林霽安：「……宋曜陽，讀書室內不准大聲喧嘩。」",
            "討論間裡，午後的陽光依舊靜靜灑落，空氣中卻多了一點活力與期待。",
            "一場不確定的相遇，就這麼靜靜地，開始了。",
        ],
    },
    # ── 新增章節範例 ─────────────────────────────
    # {
    #     "id": "story3",
    #     "title": "第三章",
    #     "subtitle": "章節副標題",
    #     "exp_threshold": 500,
    #     "requires": "story2",
    #     "excel_cell": "H5",
    #     "text": ["劇情文字1", "劇情文字2", ...],
    # },
]


# ════════════════════════════════════════════════
#  EXP 計算
#  公式：EXP = 實際分鐘數 × 完成率加成 × 專注度加成
# ════════════════════════════════════════════════

DAILY_EXP_CAP = 200

def calc_exp(actual_secs: float,
             planned_secs: float = 0,
             distraction_count: int = 0,
             daily_exp_so_far: float = 0) -> float:
    actual_min = actual_secs / 60
    if planned_secs > 0:
        completion_bonus = min(actual_secs / planned_secs, 1.2)
    else:
        completion_bonus = 1.0
    if distraction_count == 0:
        focus_bonus = 1.3
    elif distraction_count <= 2:
        focus_bonus = 1.0
    elif distraction_count <= 5:
        focus_bonus = 0.7
    else:
        focus_bonus = 0.5
    raw_exp = actual_min * completion_bonus * focus_bonus
    remaining_cap = max(0, DAILY_EXP_CAP - daily_exp_so_far)
    return min(raw_exp, remaining_cap)


# ════════════════════════════════════════════════
#  好感度計算
# ════════════════════════════════════════════════

DAILY_AFF_GAIN_CAP = 10
AFF_DECAY_PER_DAY  = 5
AFF_MIN            = 0
AFF_MAX            = 100

def calc_affection_gain(exp_gained: float, daily_aff_so_far: float = 0) -> float:
    gain = exp_gained * 0.15
    remaining_cap = max(0, DAILY_AFF_GAIN_CAP - daily_aff_so_far)
    return min(gain, remaining_cap)

def calc_affection_decay(days_absent: int) -> float:
    return AFF_DECAY_PER_DAY * min(days_absent, 4)


# ════════════════════════════════════════════════
#  AppState
# ════════════════════════════════════════════════

@dataclass
class AppState:
    start_time: float = None
    elapsed_time: float = 0
    total_time: float = 0
    average_time: float = 0
    running: bool = False
    session_count: int = 0
    user_name: str = ""
    game_data: str = "study_time.xlsx"
    story_flags: dict = field(default_factory=dict)
    total_exp: float = 0
    daily_exp: float = 0
    affection: float = 0
    daily_aff_gained: float = 0
    last_study_date: str = ""
    today_chat_shown: bool = False


# ════════════════════════════════════════════════
#  時間工具
# ════════════════════════════════════════════════

def fmt(secs: float) -> str:
    secs = max(0, int(secs))
    h = secs // 3600
    m = (secs % 3600) // 60
    s = secs % 60
    return f"{h:02}:{m:02}:{s:02}"

def fmt_min(secs: float) -> str:
    secs = max(0, int(secs))
    m = secs // 60
    s = secs % 60
    return f"{m:02}:{s:02}"

def time_str_to_sec(ts: str) -> int:
    h, m, s = map(int, ts.split(":"))
    return h * 3600 + m * 60 + s

def time_str_to_min(ts: str) -> float:
    h, m, s = map(int, ts.split(":"))
    return h * 60 + m + s / 60

def today_str() -> str:
    return datetime.date.today().strftime("%Y-%m-%d")


# ════════════════════════════════════════════════
#  Excel 存取
# ════════════════════════════════════════════════

def init_excel(game_data: str):
    if not os.path.exists(game_data):
        wb = Workbook()
        ws = wb.active
        ws.title = "紀錄"
        ws["A1"] = "總時間";      ws["B1"] = "單次讀書時間"
        ws["C1"] = "計時次數";    ws["C3"] = "平均讀書時間"
        ws["D1"] = "累積EXP";     ws["D2"] = 0
        ws["E1"] = "今日EXP";     ws["E2"] = 0
        ws["F1"] = "好感度";      ws["F2"] = 0
        ws["G1"] = "今日好感度增量"; ws["G2"] = 0
        ws["J1"] = "上次讀書日期"; ws["J2"] = ""
        ws["K1"] = "今日日常已顯示"; ws["K2"] = False
        ws["H1"] = "劇情旗標"
        ws["I1"] = "使用者名稱"
        for s in STORIES:
            ws[s["excel_cell"]] = False
        wb.save(game_data)


def load_excel_data(game_data: str) -> dict:
    wb = load_workbook(game_data)
    ws = wb["紀錄"]
    result = {
        "total_time": 0, "session_count": 0, "average_time": 0,
        "user_name": "", "story_flags": {s["id"]: False for s in STORIES},
        "total_exp": 0, "daily_exp": 0, "affection": 0,
        "daily_aff_gained": 0, "last_study_date": "", "today_chat_shown": False,
    }
    if ws["A2"].value is not None and ws["C2"].value is not None:
        result["total_time"]    = time_str_to_sec(ws["A2"].value)
        result["session_count"] = int(ws["C2"].value)
        if ws["C4"].value is not None:
            result["average_time"] = time_str_to_sec(ws["C4"].value)
    for s in STORIES:
        val = ws[s["excel_cell"]].value
        result["story_flags"][s["id"]] = bool(val) if val is not None else False
    if ws["I2"].value is not None:
        result["user_name"] = ws["I2"].value
    for cell, key, cast in [
        ("D2", "total_exp",       float),
        ("E2", "daily_exp",       float),
        ("F2", "affection",       float),
        ("G2", "daily_aff_gained",float),
        ("J2", "last_study_date", str),
        ("K2", "today_chat_shown",bool),
    ]:
        val = ws[cell].value
        if val is not None:
            result[key] = cast(val) if cast != bool else bool(val)
    return result


def save_session(game_data, total_time, elapsed_time,
                 session_count, average_time,
                 total_exp, daily_exp, affection,
                 daily_aff_gained, last_study_date):
    wb = load_workbook(game_data)
    ws = wb["紀錄"]
    row = 2
    while ws[f"B{row}"].value is not None:
        row += 1
    ws["A2"] = fmt(total_time)
    ws[f"B{row}"] = fmt(elapsed_time)
    ws["C2"] = session_count
    ws["C4"] = fmt(average_time)
    ws["D2"] = round(total_exp, 2)
    ws["E2"] = round(daily_exp, 2)
    ws["F2"] = round(affection, 2)
    ws["G2"] = round(daily_aff_gained, 2)
    ws["J2"] = last_study_date
    wb.save(game_data)


def save_story_flag(game_data, cell, value):
    wb = load_workbook(game_data)
    ws = wb["紀錄"]
    ws[cell] = value
    wb.save(game_data)


def save_user_name(game_data, name):
    wb = load_workbook(game_data)
    ws = wb["紀錄"]
    ws["I2"] = name
    wb.save(game_data)


def save_today_chat_shown(game_data):
    wb = load_workbook(game_data)
    ws = wb["紀錄"]
    ws["K2"] = True
    wb.save(game_data)


def load_session_history(game_data):
    wb = load_workbook(game_data)
    ws = wb["紀錄"]
    sessions = []
    row = 2
    while ws[f"B{row}"].value is not None:
        sessions.append(ws[f"B{row}"].value)
        row += 1
    return sessions


# ════════════════════════════════════════════════
#  好感度衰減（登入時執行）
# ════════════════════════════════════════════════

def apply_daily_decay(state: AppState) -> float:
    today = today_str()
    if state.last_study_date != today:
        state.daily_exp = 0
        state.daily_aff_gained = 0
        state.today_chat_shown = False
    if not state.last_study_date or state.last_study_date == today:
        return 0.0
    last = datetime.date.fromisoformat(state.last_study_date)
    days_absent = (datetime.date.today() - last).days - 1
    if days_absent <= 0:
        return 0.0
    decay = calc_affection_decay(days_absent)
    state.affection = max(AFF_MIN, state.affection - decay)
    return decay


# ════════════════════════════════════════════════
#  劇情觸發判定
# ════════════════════════════════════════════════

def check_story_triggers(state: AppState, on_story=None):
    for s in STORIES:
        sid = s["id"]
        if state.story_flags.get(sid):
            continue
        if s["exp_threshold"] == 0:
            continue
        req = s.get("requires")
        if req and not state.story_flags.get(req):
            continue
        if state.total_exp >= s["exp_threshold"]:
            if on_story:
                on_story(s)
            return


# ════════════════════════════════════════════════
#  計時器邏輯
# ════════════════════════════════════════════════

def start_timer(state: AppState):
    if not state.running:
        state.start_time = time.time() - state.elapsed_time
        state.running = True

def do_pause(state: AppState, callbacks: dict,
             planned_secs: float = 0, distraction_count: int = 0):
    if not state.running:
        return
    state.running = False
    actual_secs = state.elapsed_time
    exp_gained = calc_exp(actual_secs, planned_secs,
                          distraction_count, state.daily_exp)
    state.total_exp  += exp_gained
    state.daily_exp  += exp_gained
    aff_gained = calc_affection_gain(exp_gained, state.daily_aff_gained)
    state.affection = min(AFF_MAX, state.affection + aff_gained)
    state.daily_aff_gained += aff_gained
    state.total_time    += actual_secs
    state.session_count += 1
    state.average_time   = state.total_time / state.session_count
    state.last_study_date = today_str()
    save_session(state.game_data,
                 state.total_time, actual_secs,
                 state.session_count, state.average_time,
                 state.total_exp, state.daily_exp,
                 state.affection, state.daily_aff_gained,
                 state.last_study_date)
    state.elapsed_time = 0
    if callbacks.get("on_pause"):
        callbacks["on_pause"](exp_gained, aff_gained)
    check_story_triggers(state, callbacks.get("on_story"))


# ════════════════════════════════════════════════
#  劇情顯示
# ════════════════════════════════════════════════

def get_story_text(story_def, user_name):
    if "text_fn" in story_def:
        return story_def["text_fn"](user_name)
    return story_def["text"]

def show_story(story_def, user_name, story_frame, prev_frame,
               on_no_finish=None, on_finished=None):
    prev_frame.pack_forget()
    for w in story_frame.winfo_children():
        w.destroy()
    story_frame.configure(bg="#FFF9F3")
    story_frame.pack(fill="both", expand=True)
    text_list = get_story_text(story_def, user_name)
    current = {"index": 0}
    tk.Label(story_frame,
             text=f"{story_def['title']}　{story_def['subtitle']}",
             font=("Georgia", 13, "italic"),
             bg="#FFF9F3", fg=C["pink_text"]).place(x=20, y=18)
    tk.Frame(story_frame, height=1, bg=C["border"]).place(x=20, y=44, width=360)
    text_label = tk.Label(story_frame, text=text_list[0],
                          wraplength=330, font=("Georgia", 15),
                          justify="left", bg="#FFF9F3", fg="#4A1B0C",
                          padx=30, pady=20)
    text_label.place(x=0, y=60, width=400, height=480)
    prog_label = tk.Label(story_frame, text=f"1 / {len(text_list)}",
                          font=("Helvetica Neue", 11),
                          bg="#FFF9F3", fg=C["gray"])
    prog_label.place(x=0, y=560, width=400)
    hint = tk.Label(story_frame, text="點擊畫面繼續 →",
                    font=("Helvetica Neue", 11),
                    bg="#FFF9F3", fg=C["pink_light"])
    hint.place(x=0, y=585, width=400)
    def next_text(event=None):
        current["index"] += 1
        if current["index"] < len(text_list):
            text_label.config(text=text_list[current["index"]])
            prog_label.config(text=f"{current['index']+1} / {len(text_list)}")
            if on_no_finish:
                on_no_finish()
        else:
            if on_finished:
                on_finished()
            story_frame.pack_forget()
    for w in [story_frame, text_label, prog_label, hint]:
        w.bind("<Button-1>", next_text)


# ════════════════════════════════════════════════
#  UI 工具
# ════════════════════════════════════════════════

def clear_frame(frame):
    for w in frame.winfo_children():
        w.destroy()

def make_sep(parent, bg_color=None):
    return tk.Frame(parent, height=1, bg=bg_color or C["border"])


# ════════════════════════════════════════════════
#  主程式
# ════════════════════════════════════════════════

def main():
    state = AppState()
    init_excel(state.game_data)
    saved = load_excel_data(state.game_data)
    state.total_time       = saved["total_time"]
    state.session_count    = saved["session_count"]
    state.average_time     = saved["average_time"]
    state.user_name        = saved["user_name"]
    state.story_flags      = saved["story_flags"]
    state.total_exp        = saved["total_exp"]
    state.daily_exp        = saved["daily_exp"]
    state.affection        = saved["affection"]
    state.daily_aff_gained = saved["daily_aff_gained"]
    state.last_study_date  = saved["last_study_date"]
    state.today_chat_shown = saved["today_chat_shown"]

    decay_amount = apply_daily_decay(state)

    root = tk.Tk()
    root.title("讀書計時器")
    root.geometry(f"{W}x{H}")
    root.config(bg=C["bg"])
    root.resizable(False, False)

    NAV_H = 56
    content_area = tk.Frame(root, bg=C["bg"])
    content_area.place(x=0, y=0, width=W, height=H - NAV_H)

    timer_frame      = tk.Frame(content_area, bg=C["bg"])
    story_frame      = tk.Frame(content_area, bg="#FFF9F3")
    input_frame      = tk.Frame(content_area, bg=C["bg"])
    stat_frame       = tk.Frame(content_area, bg=C["bg"])
    story_list_frame = tk.Frame(content_area, bg=C["bg"])
    chat_frame       = tk.Frame(content_area, bg=C["bg"])

    ALL_FRAMES = [timer_frame, story_frame, input_frame,
                  stat_frame, story_list_frame, chat_frame]

    def show_frame(target):
        for f in ALL_FRAMES:
            f.pack_forget()
        target.pack(fill="both", expand=True)

    nav = tk.Frame(root, bg=C["surface"],
                   highlightbackground=C["border"],
                   highlightthickness=1)
    nav.place(x=0, y=H - NAV_H, width=W, height=NAV_H)

    nav_btns = {}

    def make_nav_btn(parent, label, icon, col, cmd):
        f = tk.Frame(parent, bg=C["surface"], cursor="hand2")
        f.grid(row=0, column=col, sticky="nsew")
        parent.columnconfigure(col, weight=1)
        icon_lbl = tk.Label(f, text=icon, font=("Helvetica Neue", 16),
                            bg=C["surface"], fg=C["text3"])
        icon_lbl.pack(pady=(8, 0))
        text_lbl = tk.Label(f, text=label, font=("Helvetica Neue", 9),
                            bg=C["surface"], fg=C["text3"])
        text_lbl.pack()
        for w in [f, icon_lbl, text_lbl]:
            w.bind("<Button-1>", lambda e, c=cmd: c())
        nav_btns[label] = (icon_lbl, text_lbl)

    def set_active_nav(label):
        for lbl, (icon_lbl, text_lbl) in nav_btns.items():
            if lbl == label:
                icon_lbl.config(fg=C["pink"])
                text_lbl.config(fg=C["pink"])
            else:
                icon_lbl.config(fg=C["text3"])
                text_lbl.config(fg=C["text3"])

    # ════════════════════════════════════════════
    #  頁面 1：計時器主頁
    # ════════════════════════════════════════════

    def build_timer_page():
        clear_frame(timer_frame)

        header = tk.Frame(timer_frame, bg=C["bg"])
        header.pack(fill="x", padx=20, pady=(16, 0))
        tk.Label(header, text="讀書計時器",
                 font=("Georgia", 20, "bold"),
                 bg=C["bg"], fg=C["text"]).pack(side="left")

        status_bar = tk.Frame(timer_frame, bg=C["bg"])
        status_bar.pack(fill="x", padx=20, pady=(8, 0))

        # EXP 卡
        exp_card = tk.Frame(status_bar, bg=C["gold_pale"],
                            highlightbackground="#E8C97A",
                            highlightthickness=1)
        exp_card.pack(side="left", expand=True, fill="x", padx=(0, 6))
        exp_top = tk.Frame(exp_card, bg=C["gold_pale"])
        exp_top.pack(fill="x", padx=8, pady=(6, 2))
        tk.Label(exp_top, text="EXP", font=("Helvetica Neue", 9),
                 bg=C["gold_pale"], fg=C["gold"]).pack(side="left")
        exp_val_lbl = tk.Label(exp_top, text=f"{state.total_exp:.0f}",
                               font=("Helvetica Neue", 11, "bold"),
                               bg=C["gold_pale"], fg=C["gold"])
        exp_val_lbl.pack(side="right")
        exp_bar_canvas = tk.Canvas(exp_card, height=4, bg=C["border"],
                                   highlightthickness=0)
        exp_bar_canvas.pack(fill="x", padx=8, pady=(0, 2))
        exp_next_lbl = tk.Label(exp_card, text="",
                                font=("Helvetica Neue", 8),
                                bg=C["gold_pale"], fg=C["gold"])
        exp_next_lbl.pack(pady=(0, 4))

        # 好感度卡
        aff_card = tk.Frame(status_bar, bg=C["pink_pale"],
                            highlightbackground="#F4C0D1",
                            highlightthickness=1)
        aff_card.pack(side="left", expand=True, fill="x", padx=(6, 0))
        aff_top = tk.Frame(aff_card, bg=C["pink_pale"])
        aff_top.pack(fill="x", padx=8, pady=(6, 2))
        tk.Label(aff_top, text="好感度", font=("Helvetica Neue", 9),
                 bg=C["pink_pale"], fg=C["pink_text"]).pack(side="left")
        aff_val_lbl = tk.Label(aff_top, text=f"{state.affection:.0f}",
                               font=("Helvetica Neue", 11, "bold"),
                               bg=C["pink_pale"], fg=C["pink_text"])
        aff_val_lbl.pack(side="right")
        aff_bar_canvas = tk.Canvas(aff_card, height=4, bg=C["border"],
                                   highlightthickness=0)
        aff_bar_canvas.pack(fill="x", padx=8, pady=(0, 2))
        aff_level_lbl = tk.Label(aff_card,
                                 text=get_affection_level(state.affection)["name"],
                                 font=("Helvetica Neue", 8),
                                 bg=C["pink_pale"], fg=C["pink_text"])
        aff_level_lbl.pack(pady=(0, 4))

        def _draw_bar(canvas, pct, color):
            canvas.update_idletasks()
            w = canvas.winfo_width()
            if w <= 1:
                w = 160
            canvas.delete("bar")
            fill_w = max(0, int(w * pct))
            if fill_w > 0:
                canvas.create_rectangle(0, 0, fill_w, 4,
                                        fill=color, outline="", tags="bar")

        def refresh_status_bar():
            exp_val_lbl.config(text=f"{state.total_exp:.0f}")
            aff_val_lbl.config(text=f"{state.affection:.0f}")
            aff_level_lbl.config(text=get_affection_level(state.affection)["name"])
            next_story = None
            for s in STORIES:
                if not state.story_flags.get(s["id"]) and s["exp_threshold"] > 0:
                    next_story = s
                    break
            if next_story:
                pct    = min(state.total_exp / next_story["exp_threshold"], 1.0)
                remain = max(0, next_story["exp_threshold"] - state.total_exp)
                exp_next_lbl.config(text=f"下一章還差 {remain:.0f} EXP")
            else:
                pct = 1.0
                exp_next_lbl.config(text="所有章節已解鎖！")
            _draw_bar(exp_bar_canvas, pct, C["gold_bar"])
            _draw_bar(aff_bar_canvas, state.affection / AFF_MAX, C["pink_light"])

        timer_frame.after(100, refresh_status_bar)
        refresh_status_bar()

        # 進度環
        ring_size = 200
        ring_canvas = tk.Canvas(timer_frame, width=ring_size, height=ring_size,
                                bg=C["bg"], highlightthickness=0)
        ring_canvas.pack(pady=(6, 0))
        elapsed_lbl = tk.Label(ring_canvas, text="00:00:00",
                               font=("Georgia", 24, "bold"),
                               bg=C["bg"], fg=C["text"])
        elapsed_lbl.place(relx=0.5, rely=0.40, anchor="center")
        tk.Label(ring_canvas, text="目前連續時間",
                 font=("Helvetica Neue", 10),
                 bg=C["bg"], fg=C["text3"]).place(relx=0.5, rely=0.56, anchor="center")
        total_ring_lbl = tk.Label(ring_canvas,
                                  text=f"總計 {fmt(state.total_time)}",
                                  font=("Helvetica Neue", 10),
                                  bg=C["bg"], fg=C["text2"])
        total_ring_lbl.place(relx=0.5, rely=0.68, anchor="center")
        cx, cy, r = ring_size // 2, ring_size // 2, 80

        def get_next_exp_threshold():
            for s in STORIES:
                if s["exp_threshold"] > 0 and not state.story_flags.get(s["id"]):
                    return s["exp_threshold"]
            return None

        def redraw_ring():
            ring_canvas.delete("ring")
            nxt = get_next_exp_threshold()
            pct = min(state.total_exp / nxt, 1.0) if nxt else 1.0
            ring_canvas.create_arc(cx-r, cy-r, cx+r, cy+r,
                                   start=0, extent=359.99,
                                   outline=C["border"], width=12,
                                   style="arc", tags="ring")
            if pct > 0:
                ring_canvas.create_arc(cx-r, cy-r, cx+r, cy+r,
                                       start=90, extent=-(pct*359.99),
                                       outline=C["gold_bar"], width=12,
                                       style="arc", tags="ring")
            if pct > 0.01:
                angle = math.radians(90 - pct * 360)
                tx = cx + r * math.cos(angle)
                ty = cy - r * math.sin(angle)
                ring_canvas.create_oval(tx-6, ty-6, tx+6, ty+6,
                                        fill=C["gold"], outline="", tags="ring")

        # 統計小卡列
        stat_row = tk.Frame(timer_frame, bg=C["bg"])
        stat_row.pack(fill="x", padx=20, pady=(4, 0))

        def make_mini_card(parent, label_text, value_text, value_color=None):
            card = tk.Frame(parent, bg=C["surface"],
                            highlightbackground=C["border"],
                            highlightthickness=1)
            card.pack(side="left", expand=True, fill="x", padx=3)
            tk.Label(card, text=label_text, font=("Helvetica Neue", 9),
                     bg=C["surface"], fg=C["text3"]).pack(pady=(6, 0))
            lbl = tk.Label(card, text=value_text,
                           font=("Helvetica Neue", 12, "bold"),
                           bg=C["surface"], fg=value_color or C["text"])
            lbl.pack(pady=(1, 6))
            return lbl

        count_lbl = make_mini_card(stat_row, "計時次數", f"{state.session_count} 次")
        avg_lbl   = make_mini_card(stat_row, "平均時長", fmt_min(state.average_time))
        total_lbl = make_mini_card(stat_row, "累積時間", fmt(state.total_time))

        # 按鈕列
        btn_row = tk.Frame(timer_frame, bg=C["bg"])
        btn_row.pack(fill="x", padx=20, pady=(10, 0))
        tk.Button(btn_row, text="▶  開始",
                  font=("Helvetica Neue", 13, "bold"),
                  bg=C["pink"], fg="white",
                  activebackground=C["pink_text"],
                  relief="flat", bd=0, cursor="hand2",
                  padx=0, pady=9,
                  command=lambda: start_timer(state)
                  ).pack(side="left", expand=True, fill="x", padx=(0, 5))
        tk.Button(btn_row, text="⏸  暫停",
                  font=("Helvetica Neue", 13, "bold"),
                  bg=C["surface"], fg=C["text"],
                  activebackground=C["surface2"],
                  relief="flat", bd=0, cursor="hand2",
                  highlightbackground=C["border"],
                  highlightthickness=1,
                  padx=0, pady=9,
                  command=lambda: on_pause()
                  ).pack(side="left", expand=True, fill="x", padx=(5, 0))

        # 提醒
        notice = tk.Frame(timer_frame, bg=C["pink_pale"],
                          highlightbackground="#F4C0D1",
                          highlightthickness=1)
        notice.pack(fill="x", padx=20, pady=(8, 0))
        tk.Label(notice, text="暫停後才會儲存紀錄，關閉前請記得暫停喔！",
                 font=("Helvetica Neue", 10),
                 bg=C["pink_pale"], fg=C["pink_text"],
                 wraplength=340, pady=6).pack()

        # 劇情 mini strip
        story_strip = tk.Frame(timer_frame, bg=C["bg"])
        story_strip.pack(fill="x", padx=20, pady=(8, 0))

        def refresh_story_strip():
            for w in story_strip.winfo_children():
                w.destroy()
            for s in STORIES:
                done = state.story_flags.get(s["id"], False)
                cell = tk.Frame(story_strip, bg=C["bg"])
                cell.pack(side="left", expand=True)
                c = tk.Canvas(cell, width=10, height=10,
                              bg=C["bg"], highlightthickness=0)
                c.pack()
                c.create_oval(1, 1, 9, 9,
                              fill=C["teal"] if done else C["border2"],
                              outline="")
                tk.Label(cell, text=s["title"],
                         font=("Helvetica Neue", 9),
                         bg=C["bg"], fg=C["text3"]).pack()

        refresh_story_strip()

        def update_time():
            if state.running:
                state.elapsed_time = time.time() - state.start_time
                elapsed_lbl.config(text=fmt(state.elapsed_time))
                total_ring_lbl.config(
                    text=f"總計 {fmt(state.total_time + state.elapsed_time)}")
            redraw_ring()
            root.after(1000, update_time)

        update_time()

        def on_pause():
            do_pause(state, {
                "on_pause": _after_pause,
                "on_story": trigger_story,
            }, planned_secs=0, distraction_count=0)

        def _after_pause(exp_gained, aff_gained):
            elapsed_lbl.config(text="00:00:00")
            total_ring_lbl.config(text=f"總計 {fmt(state.total_time)}")
            count_lbl.config(text=f"{state.session_count} 次")
            avg_lbl.config(text=fmt_min(state.average_time))
            total_lbl.config(text=fmt(state.total_time))
            refresh_status_bar()
            refresh_story_strip()
            toast = tk.Label(timer_frame,
                             text=f"+{exp_gained:.1f} EXP　+{aff_gained:.1f} 好感度",
                             font=("Helvetica Neue", 11, "bold"),
                             bg=C["teal_pale"], fg=C["teal_text"],
                             padx=14, pady=6,
                             highlightbackground=C["teal"],
                             highlightthickness=1)
            toast.place(relx=0.5, rely=0.97, anchor="s")
            timer_frame.after(2500, toast.destroy)

        def trigger_story(story_def):
            sid  = story_def["id"]
            cell = story_def["excel_cell"]
            def on_finished():
                state.story_flags[sid] = True
                save_story_flag(state.game_data, cell, True)
                refresh_story_strip()
                refresh_status_bar()
                show_frame(timer_frame)
            def on_no_finish():
                save_story_flag(state.game_data, cell, False)
            show_story(story_def, state.user_name,
                       story_frame, timer_frame,
                       on_no_finish=on_no_finish,
                       on_finished=on_finished)

        def trigger_story0():
            s0   = STORIES[0]
            cell = s0["excel_cell"]
            def on_finished():
                state.story_flags[s0["id"]] = True
                save_story_flag(state.game_data, cell, True)
                if not state.user_name:
                    build_input_page(back_fn=lambda: show_frame(timer_frame))
                    show_frame(input_frame)
                else:
                    show_frame(timer_frame)
            def on_no_finish():
                save_story_flag(state.game_data, cell, False)
            show_story(s0, state.user_name,
                       story_frame, timer_frame,
                       on_no_finish=on_no_finish,
                       on_finished=on_finished)

        return trigger_story, trigger_story0

    # ════════════════════════════════════════════
    #  頁面 2：輸入名字
    # ════════════════════════════════════════════

    def build_input_page(back_fn=None):
        clear_frame(input_frame)
        tk.Label(input_frame, text="請輸入你的名字",
                 font=("Georgia", 20, "bold"),
                 bg=C["bg"], fg=C["text"]).pack(pady=(80, 6))
        tk.Label(input_frame, text="名字會出現在劇情對話中",
                 font=("Helvetica Neue", 12),
                 bg=C["bg"], fg=C["text3"]).pack(pady=(0, 30))
        name_var = tk.StringVar()
        entry = tk.Entry(input_frame, textvariable=name_var,
                         font=("Helvetica Neue", 16),
                         relief="flat", bd=0,
                         highlightbackground=C["border2"],
                         highlightthickness=1,
                         bg=C["surface"], fg=C["text"],
                         insertbackground=C["text"],
                         justify="center")
        entry.pack(padx=50, ipady=10, fill="x")
        entry.focus()
        def confirm():
            name = name_var.get().strip()
            if not name:
                return
            state.user_name = name
            save_user_name(state.game_data, name)
            if back_fn:
                back_fn()
        tk.Button(input_frame, text="確認",
                  font=("Helvetica Neue", 14, "bold"),
                  bg=C["pink"], fg="white",
                  activebackground=C["pink_text"],
                  relief="flat", bd=0, cursor="hand2",
                  padx=0, pady=10,
                  command=confirm).pack(padx=50, pady=(16, 0), fill="x")
        entry.bind("<Return>", lambda e: confirm())

    # ════════════════════════════════════════════
    #  頁面 3：統計頁
    # ════════════════════════════════════════════

    def build_stat_page():
        clear_frame(stat_frame)
        tk.Label(stat_frame, text="讀書統計",
                 font=("Georgia", 18, "bold"),
                 bg=C["bg"], fg=C["text"]).pack(anchor="w", padx=20, pady=(16, 0))
        make_sep(stat_frame).pack(fill="x", padx=20, pady=(10, 0))

        grid = tk.Frame(stat_frame, bg=C["bg"])
        grid.pack(fill="x", padx=20, pady=(12, 0))

        def big_card(parent, row, col, label, value, unit=""):
            card = tk.Frame(parent, bg=C["surface"],
                            highlightbackground=C["border"],
                            highlightthickness=1)
            card.grid(row=row, column=col, padx=5, pady=5, sticky="nsew")
            tk.Label(card, text=label, font=("Helvetica Neue", 10),
                     bg=C["surface"], fg=C["text3"]).pack(pady=(10, 0))
            vf = tk.Frame(card, bg=C["surface"])
            vf.pack(pady=(2, 10))
            tk.Label(vf, text=value, font=("Georgia", 17, "bold"),
                     bg=C["surface"], fg=C["text"]).pack(side="left")
            if unit:
                tk.Label(vf, text=f" {unit}", font=("Helvetica Neue", 10),
                         bg=C["surface"], fg=C["text3"]).pack(
                             side="left", anchor="s", pady=(0, 3))

        grid.columnconfigure(0, weight=1)
        grid.columnconfigure(1, weight=1)
        h_total = int(state.total_time // 3600)
        m_total = int((state.total_time % 3600) // 60)
        total_str = f"{h_total}h {m_total:02d}m" if h_total else f"{m_total} 分"
        avg_m = int(state.average_time // 60)
        avg_s = int(state.average_time % 60)
        sessions = load_session_history(state.game_data)
        max_m = max((time_str_to_sec(t) for t in sessions), default=0) // 60
        big_card(grid, 0, 0, "累積讀書時間", total_str)
        big_card(grid, 0, 1, "平均每次時長", f"{avg_m}:{avg_s:02d}", "分鐘")
        big_card(grid, 1, 0, "計時次數", str(state.session_count), "次")
        big_card(grid, 1, 1, "最長單次", str(max_m), "分鐘")

        num_row = tk.Frame(stat_frame, bg=C["bg"])
        num_row.pack(fill="x", padx=20)

        def num_card(parent, label, value, unit, bg, fg):
            card = tk.Frame(parent, bg=bg,
                            highlightbackground=fg,
                            highlightthickness=1)
            card.pack(side="left", expand=True, fill="x", padx=5, pady=5)
            tk.Label(card, text=label, font=("Helvetica Neue", 10),
                     bg=bg, fg=fg).pack(pady=(8, 0))
            vf = tk.Frame(card, bg=bg)
            vf.pack(pady=(2, 8))
            tk.Label(vf, text=value, font=("Georgia", 17, "bold"),
                     bg=bg, fg=fg).pack(side="left")
            tk.Label(vf, text=f" {unit}", font=("Helvetica Neue", 10),
                     bg=bg, fg=fg).pack(side="left", anchor="s", pady=(0, 2))

        lvl = get_affection_level(state.affection)
        num_card(num_row, "累積 EXP", f"{state.total_exp:.0f}", "pts",
                 C["gold_pale"], C["gold"])
        num_card(num_row, f"好感度（{lvl['name']}）",
                 f"{state.affection:.0f}", "/ 100",
                 C["pink_pale"], C["pink_text"])

        tk.Label(stat_frame, text="最近 5 次讀書時長",
                 font=("Helvetica Neue", 11),
                 bg=C["bg"], fg=C["text3"]).pack(anchor="w", padx=24, pady=(10, 4))
        chart_card = tk.Frame(stat_frame, bg=C["surface"],
                              highlightbackground=C["border"],
                              highlightthickness=1)
        chart_card.pack(fill="x", padx=20)
        fig, ax = plt.subplots(figsize=(4.5, 1.8), dpi=90)
        fig.patch.set_facecolor(C["surface"])
        ax.set_facecolor(C["surface"])
        last5 = sessions[-5:] if sessions else []
        if last5:
            mins = [time_str_to_min(t) for t in last5]
            x = range(1, len(mins) + 1)
            ax.bar(x, mins,
                   color=[C["pink"] if i == len(mins) - 1
                          else C["pink_light"] for i in range(len(mins))],
                   width=0.5, zorder=3)
            ax.set_ylim(0, max(mins) * 1.3)
            ax.set_xticks(list(x))
            ax.set_xticklabels([str(i) for i in x], fontsize=9, color=C["text3"])
            ax.tick_params(axis="y", labelsize=9, labelcolor=C["text3"])
            for spine in ax.spines.values():
                spine.set_visible(False)
            ax.yaxis.set_tick_params(length=0)
            ax.xaxis.set_tick_params(length=0)
            ax.grid(axis="y", color=C["border"], linewidth=0.7, zorder=0)
        else:
            ax.text(0.5, 0.5, "還沒有紀錄", ha="center", va="center",
                    transform=ax.transAxes, color=C["text3"], fontsize=11)
            for spine in ax.spines.values():
                spine.set_visible(False)
            ax.set_xticks([]); ax.set_yticks([])
        fig.tight_layout(pad=0.8)
        canvas = FigureCanvasTkAgg(fig, master=chart_card)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="x")
        plt.close(fig)

    # ════════════════════════════════════════════
    #  頁面 4：劇情進度頁
    # ════════════════════════════════════════════

    def build_story_list_page():
        clear_frame(story_list_frame)
        tk.Label(story_list_frame, text="劇情進度",
                 font=("Georgia", 18, "bold"),
                 bg=C["bg"], fg=C["text"]).pack(anchor="w", padx=20, pady=(16, 0))
        make_sep(story_list_frame).pack(fill="x", padx=20, pady=(10, 0))

        done_count  = sum(1 for s in STORIES if state.story_flags.get(s["id"]))
        total_count = len(STORIES)
        tk.Label(story_list_frame,
                 text=f"已解鎖 {done_count} / {total_count} 章",
                 font=("Helvetica Neue", 12),
                 bg=C["bg"], fg=C["text2"]).pack(anchor="w", padx=24, pady=(8, 2))
        prog_bg = tk.Frame(story_list_frame, bg=C["border"], height=4)
        prog_bg.pack(fill="x", padx=20, pady=(0, 12))
        tk.Frame(prog_bg, bg=C["pink"], height=4).place(
            relwidth=done_count / total_count if total_count else 0, relheight=1)

        outer = tk.Frame(story_list_frame, bg=C["bg"])
        outer.pack(fill="both", expand=True, padx=20)
        cs = tk.Canvas(outer, bg=C["bg"], highlightthickness=0)
        sb = tk.Scrollbar(outer, orient="vertical", command=cs.yview)
        cs.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        cs.pack(side="left", fill="both", expand=True)
        inner = tk.Frame(cs, bg=C["bg"])
        cs.create_window((0, 0), window=inner, anchor="nw")
        inner.bind("<Configure>",
                   lambda e: cs.configure(scrollregion=cs.bbox("all")))

        for s in STORIES:
            sid  = s["id"]
            done = state.story_flags.get(sid, False)
            card = tk.Frame(inner, bg=C["surface"],
                            highlightbackground=C["border"],
                            highlightthickness=1)
            card.pack(fill="x", pady=(0, 10))
            tk.Frame(card, bg=C["teal"] if done else C["border2"],
                     width=4).pack(side="left", fill="y")
            body = tk.Frame(card, bg=C["surface"])
            body.pack(side="left", fill="both", expand=True, padx=14, pady=12)
            title_row = tk.Frame(body, bg=C["surface"])
            title_row.pack(fill="x")
            tk.Label(title_row, text=s["title"],
                     font=("Helvetica Neue", 10),
                     bg=C["surface"],
                     fg=C["teal_text"] if done else C["text3"]).pack(side="left")
            status_text = ("✓ 已解鎖" if done else
                           ("🔒 未解鎖" if s["exp_threshold"] > 0 else "🔓 可解鎖"))
            tk.Label(title_row, text=status_text,
                     font=("Helvetica Neue", 10),
                     bg=C["surface"],
                     fg=C["teal"] if done else C["gray"]).pack(side="right")
            tk.Label(body, text=s["subtitle"],
                     font=("Georgia", 14, "bold"),
                     bg=C["surface"],
                     fg=C["text"] if done else C["text3"]).pack(anchor="w", pady=(2, 0))

            if not done and s["exp_threshold"] > 0:
                need   = s["exp_threshold"]
                cur    = min(state.total_exp, need)
                pct    = cur / need
                remain = max(0, need - state.total_exp)
                tk.Label(body,
                         text=f"還差 {remain:.0f} EXP 解鎖（累積 {cur:.0f} / {need:.0f}）",
                         font=("Helvetica Neue", 10),
                         bg=C["surface"], fg=C["gold"]).pack(anchor="w", pady=(4, 2))
                bar_bg2 = tk.Frame(body, bg=C["gold_pale"], height=4)
                bar_bg2.pack(fill="x", pady=(0, 2))
                tk.Frame(bar_bg2, bg=C["gold_bar"], height=4).place(
                    relwidth=pct, relheight=1)

            req = s.get("requires")
            if not done and req and not state.story_flags.get(req):
                req_story = next((x for x in STORIES if x["id"] == req), None)
                if req_story:
                    tk.Label(body, text=f"需先解鎖「{req_story['title']}」",
                             font=("Helvetica Neue", 9),
                             bg=C["surface"], fg=C["text3"]).pack(anchor="w")

            if done:
                def make_replay(story_def=s):
                    def replay():
                        show_story(story_def, state.user_name,
                                   story_frame, story_list_frame,
                                   on_finished=lambda: show_frame(story_list_frame))
                        show_frame(story_frame)
                    return replay
                tk.Button(body, text="複習劇情",
                          font=("Helvetica Neue", 11),
                          bg=C["pink_pale"], fg=C["pink_text"],
                          activebackground="#F4C0D1",
                          relief="flat", bd=0, cursor="hand2",
                          highlightbackground="#F4C0D1",
                          highlightthickness=1,
                          padx=12, pady=5,
                          command=make_replay()).pack(anchor="w", pady=(8, 0))

        cs.bind_all("<MouseWheel>",
                    lambda e: cs.yview_scroll(int(-1*(e.delta/120)), "units"))

    # ════════════════════════════════════════════
    #  頁面 5：日常劇情聊天頁
    # ════════════════════════════════════════════

    def build_chat_page():
        import random
        clear_frame(chat_frame)

        header = tk.Frame(chat_frame, bg=C["surface"],
                          highlightbackground=C["border"],
                          highlightthickness=1)
        header.pack(fill="x")
        tk.Label(header, text="林霽安",
                 font=("Georgia", 15, "bold"),
                 bg=C["surface"], fg=C["text"]).pack(side="left", padx=16, pady=12)
        lvl = get_affection_level(state.affection)
        tk.Label(header,
                 text=f"♥ {lvl['name']}　好感度 {state.affection:.0f}",
                 font=("Helvetica Neue", 10),
                 bg=C["surface"], fg=C["pink_text"]).pack(side="right", padx=16)

        chat_outer = tk.Frame(chat_frame, bg=C["surface2"])
        chat_outer.pack(fill="both", expand=True)
        chat_canvas = tk.Canvas(chat_outer, bg=C["surface2"], highlightthickness=0)
        chat_scrollbar = tk.Scrollbar(chat_outer, orient="vertical",
                                      command=chat_canvas.yview)
        chat_canvas.configure(yscrollcommand=chat_scrollbar.set)
        chat_scrollbar.pack(side="right", fill="y")
        chat_canvas.pack(side="left", fill="both", expand=True)

        msg_area = tk.Frame(chat_canvas, bg=C["surface2"])
        msg_window = chat_canvas.create_window((0, 0), window=msg_area, anchor="nw")

        def on_msg_configure(e):
            chat_canvas.configure(scrollregion=chat_canvas.bbox("all"))
            chat_canvas.itemconfig(msg_window, width=chat_canvas.winfo_width())
        msg_area.bind("<Configure>", on_msg_configure)
        chat_canvas.bind("<Configure>",
                         lambda e: chat_canvas.itemconfig(msg_window, width=e.width))

        def add_char_bubble(text):
            row = tk.Frame(msg_area, bg=C["surface2"])
            row.pack(fill="x", padx=12, pady=(6, 0), anchor="w")
            tk.Label(row, text="霽", font=("Georgia", 11, "bold"),
                     bg=C["pink_light"], fg="white",
                     width=2, height=1, padx=4, pady=4).pack(side="left", anchor="n")
            tk.Label(row, text=text,
                     font=("Helvetica Neue", 12),
                     bg=C["bubble_char"], fg=C["text"],
                     wraplength=240, justify="left",
                     padx=12, pady=8,
                     highlightbackground=C["border"],
                     highlightthickness=1).pack(side="left", padx=(6, 0))

        def add_date_divider(text):
            row = tk.Frame(msg_area, bg=C["surface2"])
            row.pack(fill="x", padx=20, pady=(12, 4))
            tk.Frame(row, bg=C["border"], height=1).pack(side="left", expand=True, fill="x")
            tk.Label(row, text=f"  {text}  ",
                     font=("Helvetica Neue", 9),
                     bg=C["surface2"], fg=C["text3"]).pack(side="left")
            tk.Frame(row, bg=C["border"], height=1).pack(side="left", expand=True, fill="x")

        def scroll_to_bottom():
            chat_canvas.update_idletasks()
            chat_canvas.yview_moveto(1.0)

        add_date_divider("今天")

        lvl_name = get_affection_level(state.affection)["name"]
        pool = DAILY_CHATS.get(lvl_name, DAILY_CHATS["陌生"])
        random.seed(int(today_str().replace("-", "")))
        chosen = random.choice(pool)

        def show_messages(messages, idx=0):
            if idx >= len(messages):
                if not state.today_chat_shown:
                    state.today_chat_shown = True
                    save_today_chat_shown(state.game_data)
                scroll_to_bottom()
                return
            add_char_bubble(messages[idx])
            scroll_to_bottom()
            chat_frame.after(800, lambda: show_messages(messages, idx + 1))

        show_messages(chosen)

        input_bar = tk.Frame(chat_frame, bg=C["surface"],
                             highlightbackground=C["border"],
                             highlightthickness=1)
        input_bar.pack(fill="x", side="bottom")
        tk.Label(input_bar, text="選項回覆功能開發中…",
                 font=("Helvetica Neue", 11),
                 bg=C["surface"], fg=C["text3"], pady=12).pack()

    # ════════════════════════════════════════════
    #  導覽列 & 啟動
    # ════════════════════════════════════════════

    def open_timer():
        set_active_nav("計時")
        show_frame(timer_frame)

    def open_chat():
        set_active_nav("日常")
        build_chat_page()
        show_frame(chat_frame)

    def open_stat():
        if state.running:
            return
        set_active_nav("統計")
        build_stat_page()
        show_frame(stat_frame)

    def open_story_list():
        set_active_nav("劇情")
        build_story_list_page()
        show_frame(story_list_frame)

    make_nav_btn(nav, "計時", "⏱", 0, open_timer)
    make_nav_btn(nav, "日常", "💬", 1, open_chat)
    make_nav_btn(nav, "統計", "📊", 2, open_stat)
    make_nav_btn(nav, "劇情", "📖", 3, open_story_list)

    trigger_story, trigger_story0 = build_timer_page()
    set_active_nav("計時")

    if decay_amount > 0:
        def show_decay_notice():
            toast = tk.Label(content_area,
                             text=f"好感度因為你的缺席下降了 {decay_amount:.0f} 點…",
                             font=("Helvetica Neue", 11),
                             bg="#FFF0F5", fg=C["pink_text"],
                             padx=14, pady=8,
                             highlightbackground=C["pink_light"],
                             highlightthickness=1,
                             wraplength=300)
            toast.place(relx=0.5, rely=0.1, anchor="n")
            content_area.after(3000, toast.destroy)
        root.after(500, show_decay_notice)

    if state.user_name == "":
        trigger_story0()
    else:
        check_story_triggers(state, on_story=trigger_story)
        show_frame(timer_frame)

    root.mainloop()


if __name__ == "__main__":
    main()